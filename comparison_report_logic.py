import pandas as pd
import datetime
import os
import io
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from matplotlib import pyplot as plt
import tempfile
import shutil
import streamlit as st # Import Streamlit for st.error/success
from data_processing_utils import sanitize_filename # Import utility functions

def apply_comparison_filters(df_raw, comparison_config, comparison_mode, translations):
    """Applies filters and generates summary DataFrame for comparison reports."""
    years = comparison_config.get('years', [])
    months = comparison_config.get('months', [])
    selected_projects = comparison_config.get('selected_projects', [])

    df_filtered = df_raw.copy()

    if years:
        df_filtered = df_filtered[df_filtered['Year'].isin(years)]
    
    if months:
        df_filtered = df_filtered[df_filtered['MonthName'].isin(months)]
    
    if selected_projects:
        df_filtered = df_filtered[df_filtered['Project name'].isin(selected_projects)]
    else: 
        return pd.DataFrame(), translations["error_select_at_least_one_project"]

    if df_filtered.empty:
        return pd.DataFrame(), translations["warning_no_data_comparison_mode"].format(mode=comparison_mode)

    title = ""

    if comparison_mode == translations["compare_mode_project_in_month"]:
        if len(years) != 1 or len(months) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), translations["validation_project_in_month"]
        
        df_comparison = df_filtered.groupby('Project name')['Hours'].sum().reset_index()
        df_comparison.rename(columns={'Hours': translations["total_hours_col"]}, inplace=True)
        title = translations["chart_title_project_comparison_month"].format(month=months[0], year=years[0])
        return df_comparison, title

    elif comparison_mode == translations["compare_mode_project_in_year"]:
        if len(years) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), translations["validation_project_in_year"]
        
        df_comparison = df_filtered.groupby(['Project name', 'MonthName'])['Hours'].sum().unstack(fill_value=0)
        
        month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        existing_months = [m for m in month_order if m in df_comparison.columns]
        df_comparison = df_comparison[existing_months]

        df_comparison = df_comparison.reset_index().rename(columns={'index': translations["project_name_col"]})
        
        df_comparison[translations["total_hours_col"]] = df_comparison[existing_months].sum(axis=1)

        # Add total row, ensuring it's added only if there's data and not already present
        if not df_comparison.empty and 'Total' not in df_comparison[translations["project_name_col"]].values:
            total_row = df_comparison[existing_months + [translations["total_hours_col"]]].sum()
            total_row[translations["project_name_col"]] = translations["total_label"]
            df_comparison.loc[len(df_comparison)] = total_row
            
        title = translations["chart_title_project_comparison_year"].format(year=years[0])
        return df_comparison, title

    elif comparison_mode == translations["compare_mode_one_project_over_time"]:
        if len(selected_projects) != 1:
            return pd.DataFrame(), translations["error_select_only_one_project"]

        selected_project_name = selected_projects[0]

        if len(years) == 1 and len(months) > 0:
            df_comparison = df_filtered.groupby('MonthName')['Hours'].sum().reset_index()
            df_comparison.rename(columns={'Hours': translations["total_hours_for_project"].format(project=selected_project_name)}, inplace=True)
            
            month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            df_comparison['MonthName'] = pd.Categorical(df_comparison['MonthName'], categories=month_order, ordered=True)
            df_comparison = df_comparison.sort_values('MonthName').reset_index(drop=True)
            
            df_comparison[translations["project_name_col"]] = selected_project_name
            title = translations["chart_title_project_hours_over_months"].format(project=selected_project_name, year=years[0])
            return df_comparison, title

        elif len(years) > 1 and not months:
            df_comparison = df_filtered.groupby('Year')['Hours'].sum().reset_index()
            df_comparison.rename(columns={'Hours': translations["total_hours_for_project"].format(project=selected_project_name)}, inplace=True)
            df_comparison['Year'] = df_comparison['Year'].astype(str)
            
            df_comparison[translations["project_name_col"]] = selected_project_name
            title = translations["chart_title_project_hours_over_years"].format(project=selected_project_name)
            return df_comparison, title

        else:
            return pd.DataFrame(), translations["validation_one_project_over_time"]
        
    return pd.DataFrame(), translations["invalid_comparison_mode"]

def export_comparison_report_excel(df_comparison, comparison_config, output_file_path, comparison_mode, translations):
    """Exports the comparison report to an Excel file."""
    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            if df_comparison.empty:
                empty_df_for_excel = pd.DataFrame({translations["message_col"]: [translations["no_data_to_display_message"]]})
                empty_df_for_excel.to_excel(writer, sheet_name=translations["excel_comparison_sheet_name"], index=False)
            else:
                df_comparison.to_excel(writer, sheet_name=translations["excel_comparison_sheet_name"], index=False)  

            wb = writer.book
            ws = wb[translations["excel_comparison_sheet_name"]]

            data_last_row = ws.max_row
            info_row = data_last_row + 2 

            ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=4)
            ws.cell(row=info_row, column=1, value=f"{translations['comparison_report_title']}: {comparison_mode}").font = ws.cell(row=info_row, column=1).font.copy(bold=True, size=14)
            info_row += 1

            ws.cell(row=info_row, column=1, value=f"{translations['config_year_label']}:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(map(str, comparison_config.get('years', []))))
            info_row += 1
            ws.cell(row=info_row, column=1, value=f"{translations['config_months_label']}:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('months', [])))
            info_row += 1
            ws.cell(row=info_row, column=1, value=f"{translations['config_projects_label']}:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('selected_projects', [])))

            if not df_comparison.empty and len(df_comparison) > 0:
                chart = None
                data_start_row = 2 
                
                df_chart_data = df_comparison.copy()
                if translations["project_name_col"] in df_chart_data.columns and translations["total_label"] in df_chart_data[translations["project_name_col"]].values:
                    df_chart_data = df_chart_data[df_chart_data[translations["project_name_col"]] != translations["total_label"]]
                elif 'Year' in df_chart_data.columns and translations["total_label"] in df_chart_data['Year'].values:
                    df_chart_data = df_chart_data[df_chart_data['Year'] != translations["total_label"]]
                
                if df_chart_data.empty: 
                    st.warning(translations["warning_no_chart_data_excel"])
                    wb.save(output_file_path)
                    return True

                max_row_chart = data_start_row + len(df_chart_data) - 1

                if comparison_mode == translations["compare_mode_project_in_month"]:
                    chart = BarChart()
                    chart.title = translations["chart_title_project_comparison_short"]
                    chart.x_axis.title = translations["chart_xaxis_project"]
                    chart.y_axis.title = translations["chart_yaxis_hours"]
                    
                    data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(translations["total_hours_col"]) + 1, min_row=data_start_row, max_row=max_row_chart)
                    cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc(translations["project_name_col"]) + 1, min_row=data_start_row, max_row=max_row_chart) 
                    
                    chart.add_data(data_ref, titles_from_data=False) 
                    chart.set_categories(cats_ref)
                
                elif comparison_mode == translations["compare_mode_project_in_year"]:
                    chart = LineChart()
                    chart.title = translations["chart_title_project_month_comparison"]
                    chart.x_axis.title = translations["chart_xaxis_month"]
                    chart.y_axis.title = translations["chart_yaxis_hours"]

                    month_cols = [col for col in df_comparison.columns if col not in [translations["project_name_col"], translations["total_hours_col"]]]
                    
                    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
                    ordered_month_cols = [m for m in month_order if m in month_cols]

                    if ordered_month_cols:
                        min_col_month_index = df_comparison.columns.get_loc(ordered_month_cols[0])
                        max_col_month_index = df_comparison.columns.get_loc(ordered_month_cols[-1])
                        min_col_month = min_col_month_index + 1 
                        max_col_month = max_col_month_index + 1
                        cats_ref = Reference(ws, min_col=min_col_month, min_row=1, max_col=max_col_month)
                    else:
                        st.warning(translations["warning_no_month_cols_chart"])
                        wb.save(output_file_path)
                        return True
                    
                    for r_idx, project_name in enumerate(df_chart_data[translations["project_name_col"]]):
                        series_ref = Reference(ws, min_col=min_col_month, 
                                               min_row=data_start_row + r_idx, 
                                               max_col=max_col_month, 
                                               max_row=data_start_row + r_idx)
                        title_ref = Reference(ws, min_col=df_comparison.columns.get_loc(translations["project_name_col"]) + 1, 
                                              min_row=data_start_row + r_idx, 
                                              max_row=data_start_row + r_idx)
                        chart.add_data(series_ref, titles_from_data=True)
                        chart.series[r_idx].title = title_ref
                    
                    chart.set_categories(cats_ref)

                elif comparison_mode == translations["compare_mode_one_project_over_time"]:
                    total_hours_col_name = [col for col in df_comparison.columns if translations["total_hours_col_prefix"] in col][0] if [col for col in df_comparison.columns if translations["total_hours_col_prefix"] in col] else translations["total_hours_col"]
                    
                    if 'MonthName' in df_comparison.columns and len(comparison_config['years']) == 1:
                        chart = BarChart()
                        chart.title = translations["chart_title_project_over_months"].format(project=comparison_config['selected_projects'][0], year=comparison_config['years'][0])
                        chart.x_axis.title = translations["chart_xaxis_month"]
                        chart.y_axis.title = translations["chart_yaxis_hours"]
                        
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('MonthName') + 1, min_row=data_start_row, max_row=max_row_chart)
                        
                        chart.add_data(data_ref, titles_from_data=False) 
                        chart.set_categories(cats_ref)
                    elif 'Year' in df_comparison.columns and not comparison_config['months'] and len(comparison_config['years']) > 1:
                        chart = LineChart()
                        chart.title = translations["chart_title_project_over_years"].format(project=comparison_config['selected_projects'][0])
                        chart.x_axis.title = translations["chart_xaxis_year"]
                        chart.y_axis.title = translations["chart_yaxis_hours"]
                        
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Year') + 1, min_row=data_start_row, max_row=max_row_chart)
                        
                        chart.add_data(data_ref, titles_from_data=False) 
                        chart.set_categories(cats_ref)
                    else:
                        st.warning(translations["warning_invalid_time_dimension_chart"])
                        wb.save(output_file_path)
                        return True

                if chart: 
                    chart_placement_row = info_row + 2
                    ws.add_chart(chart, f"A{chart_placement_row}")

            wb.save(output_file_path)
            return True
    except Exception as e:
        st.error(translations["error_export_comparison_excel"].format(error=e))
        return False

def export_comparison_report_pdf(df_comparison, comparison_config, pdf_file_path, comparison_mode, logo_path, translations):
    """Exports the comparison PDF report with charts."""
    tmp_dir = tempfile.mkdtemp()
    charts_for_pdf = []
    today_str = datetime.datetime.today().strftime("%Y-%m-%d")

    def create_pdf_from_charts_comp(charts_data, output_path, title, config_info, logo_path_inner):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font('helvetica', 'B', 16) 

        pdf.add_page()
        if os.path.exists(logo_path_inner):
            pdf.image(logo_path_inner, x=10, y=10, w=30)
        pdf.ln(40)
        pdf.cell(0, 10, title, ln=True, align='C')
        pdf.set_font("helvetica", '', 12) 
        pdf.ln(5)
        pdf.cell(0, 10, f"{translations['generated_on_label']}: {today_str}", ln=True, align='C')
        pdf.ln(10)
        pdf.set_font("helvetica", '', 11) 
        for key, value in config_info.items():
            pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')

        for img_path, chart_title, page_project_name in charts_data:
            if img_path and os.path.exists(img_path):
                pdf.add_page()
                if os.path.exists(logo_path_inner):
                    pdf.image(logo_path_inner, x=10, y=8, w=25)
                pdf.set_font("helvetica", 'B', 11) 
                pdf.set_y(35)
                if page_project_name:
                    pdf.cell(0, 10, f"{translations['project_label']}: {page_project_name}", ln=True, align='C')
                pdf.cell(0, 10, chart_title, ln=True, align='C')
                pdf.image(img_path, x=10, y=45, w=190)

        pdf.output(output_path, "F")
        print(f"DEBUG: PDF report generated at {output_path}")

    def create_comparison_chart(df, mode, title, x_label, y_label, img_path, comparison_config_inner, translations_inner):
        fig, ax = plt.subplots(figsize=(12, 7))  
        
        df_plot = df.copy()  
        
        if translations_inner["project_name_col"] in df_plot.columns and translations_inner["total_label"] in df_plot[translations_inner["project_name_col"]].values:
            df_plot = df_plot[df_plot[translations_inner["project_name_col"]] != translations_inner["total_label"]]
        elif 'Year' in df_plot.columns and translations_inner["total_label"] in df_plot['Year'].values:
            df_plot = df_plot[df_plot['Year'] != translations_inner["total_label"]]
        
        if df_plot.empty:
            st.warning(translations_inner["warning_no_chart_data_pdf"].format(mode=mode))
            plt.close(fig)  
            return None  

        ax.set_ylim(bottom=0)
        
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'Liberation Sans']
        plt.rcParams['axes.unicode_minus'] = False 

        if mode == translations_inner["compare_mode_project_in_month"]:
            df_plot.plot(kind='bar', x=translations_inner["project_name_col"], y=translations_inner["total_hours_col"], ax=ax, color='teal')
        elif mode == translations_inner["compare_mode_project_in_year"]:
            month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            existing_months = [m for m in month_order if m in df_plot.columns]
            
            if not existing_months:
                st.warning(translations_inner["warning_no_month_cols_chart_pdf"].format(mode=mode))
                plt.close(fig)
                return None

            df_plot_long = df_plot.melt(id_vars=[translations_inner["project_name_col"]], value_vars=existing_months, var_name=translations_inner["month_label"], value_name=translations_inner["hours_label"])
            
            df_plot_long[translations_inner["month_label"]] = pd.Categorical(df_plot_long[translations_inner["month_label"]], categories=month_order, ordered=True)
            df_plot_long = df_plot_long.sort_values(translations_inner["month_label"])

            for project_name, data in df_plot_long.groupby(translations_inner["project_name_col"]):
                ax.plot(data[translations_inner["month_label"]], data[translations_inner["hours_label"]], marker='o', label=project_name)
            ax.legend(title=translations_inner["project_label"])
            ax.tick_params(axis='x', rotation=45)

        elif mode == translations_inner["compare_mode_one_project_over_time"]:
            selected_project_name = comparison_config_inner.get('selected_projects', [''])[0]
            total_hours_col_name = translations_inner["total_hours_for_project"].format(project=selected_project_name)
            
            if 'MonthName' in df_plot.columns: 
                df_plot.plot(kind='bar', x='MonthName', y=total_hours_col_name, ax=ax, color='purple')
                ax.tick_params(axis='x', rotation=45) 
            elif 'Year' in df_plot.columns: 
                df_plot.plot(kind='line', x='Year', y=total_hours_col_name, ax=ax, marker='o', color='red')
            else:
                st.warning(translations_inner["warning_invalid_time_dimension_chart_pdf"].format(mode=mode))
                plt.close(fig)
                return None
        else:
            st.warning(translations_inner["warning_unknown_comparison_mode_chart"].format(mode=mode))
            plt.close(fig)
            return None

        ax.set_title(title, fontsize=12)
        ax.set_xlabel(x_label, fontsize=10)
        ax.set_ylabel(y_label, fontsize=10)
        
        plt.tight_layout()
        fig.savefig(img_path, dpi=200)
        plt.close(fig)
        return img_path

    try:
        pdf_config_info = {
            translations["comparison_mode_label"]: comparison_mode,
            translations["config_years_label"]: ', '.join(map(str, comparison_config.get('years', []))) if comparison_config.get('years') else "N/A",
            translations["config_months_label"]: ', '.join(comparison_config.get('months', [])) if comparison_config.get('months') else translations["all_label"],
            translations["config_projects_label"]: ', '.join(comparison_config.get('selected_projects', [])) if comparison_config.get('selected_projects') else translations["none_label"]
        }

        main_chart_path = None
        chart_title = ""
        x_label = ""
        y_label = translations["hours_label"]
        page_project_name_for_chart = None

        if comparison_mode == translations["compare_mode_project_in_month"]:
            chart_title = translations["chart_title_project_comparison_month"].format(month=comparison_config['months'][0], year=comparison_config['years'][0])
            x_label = translations["chart_xaxis_project"]
            main_chart_path = create_comparison_chart(df_comparison, comparison_mode, chart_title, x_label, y_label, 
                                                     os.path.join(tmp_dir, "comparison_chart_month.png"), comparison_config, translations)
            if main_chart_path: charts_for_pdf.append((main_chart_path, chart_title, None))


        elif comparison_mode == translations["compare_mode_project_in_year"]:
            chart_title = translations["chart_title_project_comparison_year"].format(year=comparison_config['years'][0])
            x_label = translations["chart_xaxis_month"]
            main_chart_path = create_comparison_chart(df_comparison, comparison_mode, chart_title, x_label, y_label, 
                                                     os.path.join(tmp_dir, "comparison_chart_year.png"), comparison_config, translations)
            if main_chart_path: charts_for_pdf.append((main_chart_path, chart_title, None))
            
        elif comparison_mode == translations["compare_mode_one_project_over_time"]:
            selected_proj = comparison_config.get('selected_projects', [''])[0]
            page_project_name_for_chart = selected_proj

            if len(comparison_config.get('years', [])) == 1 and len(comparison_config.get('months', [])) > 0:
                chart_title = translations["chart_title_project_hours_over_months"].format(project=selected_proj, year=comparison_config['years'][0])
                x_label = translations["chart_xaxis_month"]
                main_chart_path = create_comparison_chart(df_comparison, comparison_mode, chart_title, x_label, y_label,
                                                         os.path.join(tmp_dir, f"{sanitize_filename(selected_proj)}_months_chart.png"), comparison_config, translations)
            elif len(comparison_config.get('years', [])) > 1 and not comparison_config.get('months', []):
                chart_title = translations["chart_title_project_hours_over_years"].format(project=selected_proj)
                x_label = translations["chart_xaxis_year"]
                main_chart_path = create_comparison_chart(df_comparison, comparison_mode, chart_title, x_label, y_label,
                                                         os.path.join(tmp_dir, f"{sanitize_filename(selected_proj)}_years_chart.png"), comparison_config, translations)
            else:
                st.warning(translations["warning_invalid_time_dimension_chart_pdf_config"])
                main_chart_path = None
            
            if main_chart_path:
                charts_for_pdf.append((main_chart_path, chart_title, page_project_name_for_chart))

        if not charts_for_pdf:
            st.warning(translations["warning_no_charts_comparison_pdf"])
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, translations["pdf_report_title_comparison"], ln=True, align='C')
            pdf.set_font("helvetica", '', 12)
            pdf.cell(0, 10, f"{translations['generated_on_label']}: {today_str}", ln=True, align='C')
            pdf.ln(10)
            pdf.set_font("helvetica", '', 11)
            for key, value in pdf_config_info.items():
                pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')
            pdf.cell(0, 10, translations["no_charts_generated_message"], ln=True, align='C')
            pdf.output(pdf_file_path, "F")
            return True

        create_pdf_from_charts_comp(charts_for_pdf, pdf_file_path, translations["pdf_report_title_comparison"], pdf_config_info, logo_path)
        return True

    except Exception as e:
        st.error(translations["error_create_comparison_pdf"].format(error=e))
        return False
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)
