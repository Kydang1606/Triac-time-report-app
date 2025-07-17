import pandas as pd
import datetime
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from matplotlib import pyplot as plt
import tempfile
import re
import shutil

# --- Hàm hỗ trợ ---
def sanitize_filename(name):
    """
    Làm sạch một chuỗi để sử dụng làm tên tệp hoặc tên sheet Excel hợp lệ.
    Loại bỏ các ký tự không hợp lệ và giới hạn độ dài 31 ký tự cho sheet Excel.
    """
    invalid_chars = re.compile(r'[\\/*?[\]:;|=,<>]')
    s = invalid_chars.sub("_", str(name))
    s = ''.join(c for c in s if c.isprintable())
    return s[:31]

def setup_paths():
    """
    Thiết lập và trả về từ điển các đường dẫn tệp đầu vào và đầu ra.
    """
    today = datetime.datetime.today().strftime('%Y%m%d')
    return {
        'template_file': "Time_report.xlsm",
        'output_file': f"Time_report_Standard_{today}.xlsx",
        'pdf_report': f"Time_report_Standard_{today}.pdf",
        'comparison_output_file': f"Time_report_Comparison_{today}.xlsx",
        'comparison_pdf_report': f"Time_report_Comparison_{today}.pdf",
        'logo_path': "triac_logo.png"
    }

def read_configs(template_file):
    """
    Đọc các cài đặt cấu hình (chế độ, năm, tháng, bộ lọc dự án)
    từ tệp Excel mẫu được chỉ định.
    """
    try:
        year_mode_df = pd.read_excel(template_file, sheet_name='Config_Year_Mode', engine='openpyxl')
        project_filter_df = pd.read_excel(template_file, sheet_name='Config_Project_Filter', engine='openpyxl')

        mode_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'mode', 'Value']
        mode = str(mode_row.values[0]).strip().lower() if not mode_row.empty and pd.notna(mode_row.values[0]) else 'year'

        year_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'year', 'Value']
        year = int(year_row.values[0]) if not year_row.empty and pd.notna(year_row.values[0]) and pd.api.types.is_number(year_row.values[0]) else datetime.datetime.now().year
        
        months_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'months', 'Value']
        months = [m.strip().capitalize() for m in str(months_row.values[0]).split(',')] if not months_row.empty and pd.notna(months_row.values[0]) else []
        
        if 'Include' in project_filter_df.columns:
            project_filter_df['Include'] = project_filter_df['Include'].astype(str).str.lower()

        return {
            'mode': mode,
            'year': year,
            'months': months,
            'project_filter_df': project_filter_df
        }
    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy tệp mẫu tại {template_file}")
        return {'mode': 'year', 'year': datetime.datetime.now().year, 'months': [], 'project_filter_df': pd.DataFrame(columns=['Project Name', 'Include'])}
    except Exception as e:
        print(f"Lỗi khi đọc cấu hình: {e}")
        return {'mode': 'year', 'year': datetime.datetime.now().year, 'months': [], 'project_filter_df': pd.DataFrame(columns=['Project Name', 'Include'])}

def load_raw_data(template_file):
    """
    Tải dữ liệu thô từ sheet 'Raw Data' của tệp Excel mẫu được chỉ định.
    Thực hiện làm sạch dữ liệu ban đầu và chuẩn bị các cột ngày cần thiết.
    """
    try:
        df = pd.read_excel(template_file, sheet_name='Raw Data', engine='openpyxl')
        df.columns = df.columns.str.strip()
        df.rename(columns={'Hou': 'Hours', 'Team member': 'Employee', 'Project Name': 'Project name'}, inplace=True)
        
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date'])
        
        df['Year'] = df['Date'].dt.year
        df['MonthName'] = df['Date'].dt.month_name()
        df['Week'] = df['Date'].dt.isocalendar().week.astype(int)
        
        df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce').fillna(0)
        
        return df
    except Exception as e:
        print(f"Lỗi khi tải dữ liệu thô: {e}")
        return pd.DataFrame()

def apply_filters(df, config):
    """
    Áp dụng các bộ lọc vào DataFrame dựa trên cấu hình được cung cấp.
    Lọc theo năm, tháng và các dự án đã chọn.
    """
    df_filtered = df.copy()

    if 'years' in config and config['years']:
        df_filtered = df_filtered[df_filtered['Year'].isin(config['years'])]
    elif 'year' in config and config['year']:
        df_filtered = df_filtered[df_filtered['Year'] == config['year']]

    if config['months']:
        df_filtered = df_filtered[df_filtered['MonthName'].isin(config['months'])]

    if not config['project_filter_df'].empty:
        selected_project_names = config['project_filter_df']['Project Name'].tolist()
        df_filtered = df_filtered[df_filtered['Project name'].isin(selected_project_names)]
    else:
        return pd.DataFrame(columns=df.columns)

    return df_filtered

def export_report(df, config, output_file_path):
    """
    Xuất báo cáo tiêu chuẩn ra tệp Excel.
    Bao gồm một sheet tổng quan, các sheet dự án riêng lẻ và biểu đồ.
    """
    mode = config.get('mode', 'year')
    
    groupby_cols = []
    if mode == 'year':
        groupby_cols = ['Year', 'Project name']
    elif mode == 'month':
        groupby_cols = ['Year', 'MonthName', 'Project name']
    else: # week mode
        groupby_cols = ['Year', 'Week', 'Project name']

    for col in groupby_cols + ['Hours']:
        if col not in df.columns:
            print(f"Lỗi: Cột '{col}' không tồn tại trong DataFrame. Không thể tạo báo cáo.")
            return False

    if df.empty:
        print("Cảnh báo: DataFrame đã lọc trống, không có báo cáo nào được tạo.")
        return False

    summary = df.groupby(groupby_cols)['Hours'].sum().reset_index()

    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            summary.to_excel(writer, sheet_name='Summary', index=False)

        wb = load_workbook(output_file_path)
        ws = wb['Summary']
        
        if len(summary) > 0:
            data_col_idx = summary.columns.get_loc('Hours') + 1
            cats_col_idx = summary.columns.get_loc('Project name') + 1

            data_ref = Reference(ws, min_col=data_col_idx, min_row=2, max_row=ws.max_row)
            cats_ref = Reference(ws, min_col=cats_col_idx, min_row=2, max_row=ws.max_row)

            chart = BarChart()
            chart.title = f"Tổng giờ theo dự án ({mode})"
            chart.x_axis.title = "Dự án"
            chart.y_axis.title = "Giờ"
            
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "F2")

        for project in df['Project name'].unique():
            df_proj = df[df['Project name'] == project]
            sheet_title = sanitize_filename(project)
            
            if sheet_title in wb.sheetnames:
                ws_proj = wb[sheet_title]
            else:
                ws_proj = wb.create_sheet(title=sheet_title)

            summary_task = df_proj.groupby('Task')['Hours'].sum().reset_index().sort_values('Hours', ascending=False)
            
            if not summary_task.empty:
                ws_proj.append(['Task', 'Hours'])
                for row_data in dataframe_to_rows(summary_task, index=False, header=False):
                    ws_proj.append(row_data)

                chart_task = BarChart()
                chart_task.title = f"{project} - Giờ theo công việc"
                chart_task.x_axis.title = "Công việc"
                chart_task.y_axis.title = "Giờ"
                task_len = len(summary_task)
                
                data_ref_task = Reference(ws_proj, min_col=2, min_row=1, max_row=task_len + 1)
                cats_ref_task = Reference(ws_proj, min_col=1, min_row=2, max_row=task_len + 1)
                chart_task.add_data(data_ref_task, titles_from_data=True)
                chart_task.set_categories(cats_ref_task)
                ws_proj.add_chart(chart_task, f"E1")

            start_row_raw_data = ws_proj.max_row + 2 if ws_proj.max_row > 1 else 1
            if not summary_task.empty:
                start_row_raw_data += 15 # Di chuyển dữ liệu thô xuống xa hơn nếu có bảng tổng quan công việc

            for r_idx, r in enumerate(dataframe_to_rows(df_proj, index=False, header=True)):
                for c_idx, cell_val in enumerate(r):
                    ws_proj.cell(row=start_row_raw_data + r_idx, column=c_idx + 1, value=cell_val)
        
        ws_config = wb.create_sheet("Config_Info")
        ws_config['A1'], ws_config['B1'] = "Chế độ", config.get('mode', 'N/A').capitalize()
        ws_config['A2'], ws_config['B2'] = "Năm", ', '.join(map(str, config.get('years', []))) if config.get('years') else str(config.get('year', 'N/A'))
        ws_config['A3'], ws_config['B3'] = "Tháng", ', '.join(config.get('months', [])) if config.get('months') else "Tất cả"
        
        if 'project_filter_df' in config and not config['project_filter_df'].empty:
            selected_projects_display = config['project_filter_df'][config['project_filter_df']['Include'].astype(str).str.lower() == 'yes']['Project Name'].tolist()
            ws_config['A4'], ws_config['B4'] = "Dự án bao gồm", ', '.join(selected_projects_display)
        else:
            ws_config['A4'], ws_config['B4'] = "Dự án bao gồm", "Không có dự án nào được chọn hoặc tìm thấy"

        # Xóa các sheet mẫu
        for sheet_name in ['Raw Data', 'Config_Year_Mode', 'Config_Project_Filter']:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        wb.save(output_file_path)
        return True
    except Exception as e:
        print(f"Lỗi khi xuất báo cáo tiêu chuẩn: {e}")
        return False

def export_pdf_report(df, config, pdf_report_path, logo_path):
    """
    Xuất báo cáo PDF tiêu chuẩn với biểu đồ cho Workcentre và Giờ theo công việc của mỗi dự án.
    """
    today_str = datetime.datetime.today().strftime("%Y-%m-%d")
    tmp_dir = tempfile.mkdtemp()
    charts_for_pdf = []

    def create_pdf_from_charts(charts_data, output_path, title, config_info, logo_path_inner):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font('helvetica', 'B', 16)
        
        pdf.add_page()
        if os.path.exists(logo_path_inner):
            pdf.image(logo_path_inner, x=10, y=10, w=30)
        pdf.ln(40)
        pdf.cell(0, 10, title, ln=True, align='C')
        pdf.set_font("helvetica", '', 12)
        pdf.ln(5)
        pdf.cell(0, 10, f"Được tạo vào: {today_str}", ln=True, align='C')
        pdf.ln(10)
        pdf.set_font("helvetica", '', 11)
        for key, value in config_info.items():
            pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')

        for img_path, chart_title, page_project_name in charts_data:
            if img_path and os.path.exists(img_path):
                pdf.add_page()
                if os.path.exists(logo_path_inner):
                    pdf.image(logo_path_inner, x=10, y=8, w=25)
                pdf.set_font("helvetica", 'B', 11)
                pdf.set_y(35)
                if page_project_name:
                    pdf.cell(0, 10, f"Dự án: {page_project_name}", ln=True, align='C')
                pdf.cell(0, 10, chart_title, ln=True, align='C')
                pdf.image(img_path, x=10, y=45, w=190)

        pdf.output(output_path, "F")
        print(f"DEBUG: Báo cáo PDF được tạo tại {output_path}")

    try:
        projects = df['Project name'].unique()

        config_info = {
            "Chế độ": config.get('mode', 'N/A').capitalize(),
            "Năm": ', '.join(map(str, config.get('years', []))) if config.get('years') else str(config.get('year', 'N/A')),
            "Tháng": ', '.join(config.get('months', [])) if config.get('months') else "Tất cả",
            "Dự án bao gồm": ', '.join(config['project_filter_df']['Project Name']) if 'project_filter_df' in config and not config['project_filter_df'].empty else "Không có dự án nào được chọn hoặc tìm thấy"
        }

        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'Liberation Sans']
        plt.rcParams['axes.unicode_minus'] = False

        for project in projects:
            safe_project = sanitize_filename(project)
            df_proj = df[df['Project name'] == project]

            if 'Workcentre' in df_proj.columns and not df_proj['Workcentre'].empty:
                workcentre_summary = df_proj.groupby('Workcentre')['Hours'].sum().sort_values(ascending=False)
                if not workcentre_summary.empty and workcentre_summary.sum() > 0:
                    fig, ax = plt.subplots(figsize=(10, 5))
                    workcentre_summary.plot(kind='barh', color='skyblue', ax=ax)
                    ax.set_title(f"{project} - Giờ theo Workcentre", fontsize=9)
                    ax.tick_params(axis='y', labelsize=8)
                    ax.set_xlabel("Giờ")
                    ax.set_ylabel("Workcentre")
                    wc_img_path = os.path.join(tmp_dir, f"{safe_project}_wc.png")
                    plt.tight_layout()
                    fig.savefig(wc_img_path, dpi=150)
                    plt.close(fig)
                    charts_for_pdf.append((wc_img_path, f"{project} - Giờ theo Workcentre", project))

            if 'Task' in df_proj.columns and not df_proj['Task'].empty:
                task_summary = df_proj.groupby('Task')['Hours'].sum().sort_values(ascending=False)
                if not task_summary.empty and task_summary.sum() > 0:
                    fig, ax = plt.subplots(figsize=(10, 6))
                    task_summary.plot(kind='barh', color='lightgreen', ax=ax)
                    ax.set_title(f"{project} - Giờ theo công việc", fontsize=9)
                    ax.tick_params(axis='y', labelsize=8)
                    ax.set_xlabel("Giờ")
                    ax.set_ylabel("Công việc")
                    task_img_path = os.path.join(tmp_dir, f"{safe_project}_task.png")
                    plt.tight_layout()
                    fig.savefig(task_img_path, dpi=150)
                    plt.close(fig)
                    charts_for_pdf.append((task_img_path, f"{project} - Giờ theo công việc", project))

        if not charts_for_pdf:
            print("Cảnh báo: Không có biểu đồ nào được tạo để đưa vào PDF. PDF có thể trống.")
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, "BÁO CÁO GIỜ TRIAC - TIÊU CHUẨN", ln=True, align='C')
            pdf.set_font("helvetica", '', 12)
            pdf.cell(0, 10, f"Được tạo vào: {today_str}", ln=True, align='C')
            pdf.ln(10)
            pdf.set_font("helvetica", '', 11)
            for key, value in config_info.items():
                pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')
            pdf.cell(0, 10, "Không có biểu đồ nào được tạo cho báo cáo này.", ln=True, align='C')
            pdf.output(pdf_report_path, "F")
            return True
            
        create_pdf_from_charts(charts_for_pdf, pdf_report_path, "BÁO CÁO GIỜ TRIAC - TIÊU CHUẨN", config_info, logo_path)
        return True
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo PDF: {e}")
        return False
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)

def apply_comparison_filters(df_raw, comparison_config, comparison_mode):
    """
    Áp dụng bộ lọc và tạo DataFrame tóm tắt cho báo cáo so sánh.
    Xử lý các chế độ so sánh khác nhau (dự án trong một tháng/năm, một dự án theo thời gian).
    """
    print("DEBUG: apply_comparison_filters được gọi với:")
    print(f"  kiểu df_raw: {type(df_raw)}")
    print(f"  kiểu comparison_config: {type(comparison_config)}")
    print(f"  kiểu comparison_mode: {type(comparison_mode)} giá trị: {comparison_mode}")

    years = comparison_config.get('years', [])
    months = comparison_config.get('months', [])
    selected_projects = comparison_config.get('selected_projects', [])

    df_filtered = df_raw.copy()

    if years:
        df_filtered = df_filtered[df_filtered['Year'].isin(years)]
    
    if months:
        df_filtered = df_filtered[df_filtered['MonthName'].isin(months)]
    
    if selected_projects:
        df_filtered = df_filtered[df_filtered['Project name'].isin(selected_projects)]
    else:
        return pd.DataFrame(), "Vui lòng chọn ít nhất một dự án để so sánh."

    if df_filtered.empty:
        return pd.DataFrame(), f"Không tìm thấy dữ liệu cho chế độ so sánh: {comparison_mode} với các lựa chọn hiện tại."

    title = ""

    if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
        if len(years) != 1 or len(months) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), "Vui lòng chọn MỘT năm, MỘT tháng và ít nhất HAI dự án cho chế độ này."
        
        df_comparison = df_filtered.groupby('Project name')['Hours'].sum().reset_index()
        df_comparison.rename(columns={'Hours': 'Total Hours'}, inplace=True)
        title = f"So sánh giờ giữa các dự án trong {months[0]}, năm {years[0]}"
        return df_comparison, title

    elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
        if len(years) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), "Vui lòng chọn MỘT năm và ít nhất HAI dự án cho chế độ này."
        
        df_comparison = df_filtered.groupby(['Project name', 'MonthName'])['Hours'].sum().unstack(fill_value=0)
        
        month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        existing_months = [m for m in month_order if m in df_comparison.columns]
        df_comparison = df_comparison[existing_months]

        df_comparison = df_comparison.reset_index().rename(columns={'index': 'Project Name'})
        
        df_comparison['Total Hours'] = df_comparison[existing_months].sum(axis=1)

        df_comparison.loc['Total'] = df_comparison[existing_months + ['Total Hours']].sum()
        df_comparison.loc['Total', 'Project Name'] = 'Total'

        title = f"So sánh giờ giữa các dự án trong năm {years[0]} (theo tháng)"
        return df_comparison, title

    elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
        if len(selected_projects) != 1:
            return pd.DataFrame(), "Lỗi: Nội bộ - Vui lòng chọn CHỈ MỘT dự án cho chế độ này."

        selected_project_name = selected_projects[0]

        if len(years) == 1 and len(months) > 0:
            df_comparison = df_filtered.groupby('MonthName')['Hours'].sum().reset_index()
            df_comparison.rename(columns={'Hours': f'Total Hours for {selected_project_name}'}, inplace=True)
            
            month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            df_comparison['MonthName'] = pd.Categorical(df_comparison['MonthName'], categories=month_order, ordered=True)
            df_comparison = df_comparison.sort_values('MonthName').reset_index(drop=True)
            
            df_comparison['Project Name'] = selected_project_name
            title = f"Tổng giờ dự án {selected_project_name} qua các tháng trong năm {years[0]}"
            return df_comparison, title

        elif len(years) > 1 and not months:
            df_comparison = df_filtered.groupby('Year')['Hours'].sum().reset_index()
            df_comparison.rename(columns={'Hours': f'Total Hours for {selected_project_name}'}, inplace=True)
            df_comparison['Year'] = df_comparison['Year'].astype(str)
            
            df_comparison['Project Name'] = selected_project_name
            title = f"Tổng giờ dự án {selected_project_name} qua các năm"
            return df_comparison, title

        else:
            return pd.DataFrame(), "Cấu hình so sánh dự án qua thời gian không hợp lệ. Vui lòng chọn một năm với nhiều tháng, HOẶC nhiều năm."
        
    return pd.DataFrame(), "Chế độ so sánh không hợp lệ."


def export_comparison_report(df_comparison, comparison_config, output_file_path, comparison_mode):
    """
    Xuất báo cáo so sánh ra tệp Excel.
    """
    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            if df_comparison.empty:
                empty_df_for_excel = pd.DataFrame({"Thông báo": ["Không có dữ liệu để hiển thị với các bộ lọc đã chọn."]})
                empty_df_for_excel.to_excel(writer, sheet_name='Comparison Report', index=False)
            else:
                df_comparison.to_excel(writer, sheet_name='Comparison Report', index=False)

            wb = writer.book
            ws = wb['Comparison Report']

            data_last_row = ws.max_row
            info_row = data_last_row + 2

            ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=4)
            ws.cell(row=info_row, column=1, value=f"BÁO CÁO SO SÁNH: {comparison_mode}").font = ws.cell(row=info_row, column=1).font.copy(bold=True, size=14)
            info_row += 1

            ws.cell(row=info_row, column=1, value="Năm:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(map(str, comparison_config.get('years', []))))
            info_row += 1
            ws.cell(row=info_row, column=1, value="Tháng:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('months', [])))
            info_row += 1
            ws.cell(row=info_row, column=1, value="Dự án:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('selected_projects', [])))

            if not df_comparison.empty and len(df_comparison) > 0:
                chart = None
                data_start_row = 2
                
                df_chart_data = df_comparison.copy()
                if 'Project Name' in df_chart_data.columns and 'Total' in df_chart_data['Project Name'].values:
                    df_chart_data = df_chart_data[df_chart_data['Project Name'] != 'Total']
                elif 'Year' in df_chart_data.columns and 'Total' in df_chart_data['Year'].values:
                    df_chart_data = df_chart_data[df_chart_data['Year'] != 'Total']
                
                if df_chart_data.empty:
                    print("Không đủ dữ liệu để vẽ biểu đồ so sánh sau khi loại bỏ hàng tổng.")
                    wb.save(output_file_path)
                    return True

                max_row_chart = data_start_row + len(df_chart_data) - 1

                if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
                    chart = BarChart()
                    chart.title = "So sánh giờ theo dự án"
                    chart.x_axis.title = "Dự án"
                    chart.y_axis.title = "Giờ"
                    
                    data_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Total Hours') + 1, min_row=data_start_row, max_row=max_row_chart)
                    cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Project name') + 1, min_row=data_start_row, max_row=max_row_chart)
                    
                    chart.add_data(data_ref, titles_from_data=False)
                    chart.set_categories(cats_ref)
                
                elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
                    chart = LineChart()
                    chart.title = "So sánh giờ theo dự án và tháng"
                    chart.x_axis.title = "Tháng"
                    chart.y_axis.title = "Giờ"

                    month_cols = [col for col in df_comparison.columns if col not in ['Project Name', 'Total Hours']]
                    
                    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
                    ordered_month_cols = [m for m in month_order if m in month_cols]

                    if ordered_month_cols:
                        min_col_month_index = df_comparison.columns.get_loc(ordered_month_cols[0])
                        max_col_month_index = df_comparison.columns.get_loc(ordered_month_cols[-1])
                        min_col_month = min_col_month_index + 1
                        max_col_month = max_col_month_index + 1
                        cats_ref = Reference(ws, min_col=min_col_month, min_row=1, max_col=max_col_month)
                    else:
                        print("Không tìm thấy cột tháng để tạo biểu đồ.")
                        wb.save(output_file_path)
                        return True
                    
                    for r_idx, project_name in enumerate(df_chart_data['Project Name']):
                        series_ref = Reference(ws, min_col=min_col_month,
                                                min_row=data_start_row + r_idx,
                                                max_col=max_col_month,
                                                max_row=data_start_row + r_idx)
                        title_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Project Name') + 1,
                                                min_row=data_start_row + r_idx,
                                                max_row=data_start_row + r_idx)
                        chart.add_data(series_ref, titles_from_data=True)
                        chart.series[r_idx].title = title_ref
                    
                    chart.set_categories(cats_ref)

                elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
                    total_hours_col_name = [col for col in df_comparison.columns if 'Total Hours' in col][0] if [col for col in df_comparison.columns if 'Total Hours' in col] else 'Total Hours'
                    
                    if 'MonthName' in df_comparison.columns and len(comparison_config['years']) == 1:
                        chart = BarChart()
                        chart.title = f"Tổng giờ dự án {comparison_config['selected_projects'][0]} năm {comparison_config['years'][0]} theo tháng"
                        chart.x_axis.title = "Tháng"
                        chart.y_axis.title = "Giờ"
                        
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('MonthName') + 1, min_row=data_start_row, max_row=max_row_chart)
                        
                        chart.add_data(data_ref, titles_from_data=False)
                        chart.set_categories(cats_ref)
                    elif 'Year' in df_comparison.columns and not comparison_config['months'] and len(comparison_config['years']) > 1:
                        chart = LineChart()
                        chart.title = f"Tổng giờ dự án {comparison_config['selected_projects'][0]} qua các năm"
                        chart.x_axis.title = "Năm"
                        chart.y_axis.title = "Giờ"
                        
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Year') + 1, min_row=data_start_row, max_row=max_row_chart)
                        
                        chart.add_data(data_ref, titles_from_data=False)
                        chart.set_categories(cats_ref)
                    else:
                        raise ValueError("Không tìm thấy kích thước thời gian hợp lệ cho các danh mục biểu đồ trong chế độ so sánh qua tháng/năm.")

                if chart:
                    chart_placement_row = info_row + 2
                    ws.add_chart(chart, f"A{chart_placement_row}")

            wb.save(output_file_path)
            return True
    except Exception as e:
        print(f"Lỗi khi xuất báo cáo so sánh ra Excel: {e}")
        return False

def export_comparison_pdf_report(df_comparison, comparison_config, pdf_file_path, comparison_mode, logo_path):
    """
    Xuất báo cáo PDF so sánh với biểu đồ.
    """
    tmp_dir = tempfile.mkdtemp()
    charts_for_pdf = []

    def create_pdf_from_charts_comp(charts_data, output_path, title, config_info, logo_path_inner):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font('helvetica', 'B', 16)

        pdf.add_page()
        if os.path.exists(logo_path_inner):
            pdf.image(logo_path_inner, x=10, y=10, w=30)
        pdf.ln(40)
        pdf.cell(0, 10, title, ln=True, align='C')
        pdf.set_font("helvetica", '', 12)
        pdf.ln(5)
        pdf.cell(0, 10, f"Được tạo vào: {datetime.datetime.today().strftime('%Y-%m-%d')}", ln=True, align='C')
        pdf.ln(10)
        pdf.set_font("helvetica", '', 11)
        for key, value in config_info.items():
            pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')

        for img_path, chart_title, page_project_name in charts_data:
            if img_path and os.path.exists(img_path):
                pdf.add_page()
                if os.path.exists(logo_path_inner):
                    pdf.image(logo_path_inner, x=10, y=8, w=25)
                pdf.set_font("helvetica", 'B', 11)
                pdf.set_y(35)
                if page_project_name:
                    pdf.cell(0, 10, f"Dự án: {page_project_name}", ln=True, align='C')
                pdf.cell(0, 10, chart_title, ln=True, align='C')
                pdf.image(img_path, x=10, y=45, w=190)

        pdf.output(output_path, "F")
        print(f"DEBUG: Báo cáo PDF được tạo tại {output_path}")

    def create_comparison_chart(df, mode, title, x_label, y_label, img_path, comparison_config_inner):
        fig, ax = plt.subplots(figsize=(12, 7))
        
        df_plot = df.copy()
        
        if 'Project Name' in df_plot.columns and 'Total' in df_plot['Project Name'].values:
            df_plot = df_plot[df_plot['Project Name'] != 'Total']
        elif 'Year' in df_plot.columns and 'Total' in df_plot['Year'].values:
            df_plot = df_plot[df_plot['Year'] != 'Total']
        
        if df_plot.empty:
            print(f"DEBUG: df_plot trống cho chế độ '{mode}' sau khi bỏ hàng 'Total'. Bỏ qua việc tạo biểu đồ.")
            plt.close(fig)
            return None

        ax.set_ylim(bottom=0)
        
        # Đặt lại font cho biểu đồ matplotlib để hiển thị tiếng Việt nếu cần
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'Helvetica', 'Liberation Sans'] # Thêm font hỗ trợ tiếng Việt
        plt.rcParams['axes.unicode_minus'] = False # Giúp hiển thị dấu trừ đúng

        if mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
            df_plot.plot(kind='bar', x='Project name', y='Total Hours', ax=ax, color='purple')
            ax.set_xticklabels(df_plot['Project name'], rotation=45, ha="right", fontsize=10)
        elif mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
            # Bỏ cột 'Total Hours' trước khi vẽ biểu đồ vì nó không phải là tháng
            plot_df = df_plot.drop(columns=['Total Hours'], errors='ignore')
            plot_df.set_index('Project Name').T.plot(kind='line', marker='o', ax=ax)
            ax.set_xticks(range(len(plot_df.columns) - 1)) # Loại trừ Project Name
            month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            # Lọc các tháng thực tế có trong cột và sắp xếp chúng
            ordered_months = [m for m in month_order if m in plot_df.columns]
            ax.set_xticklabels(ordered_months, rotation=45, ha="right", fontsize=10)
            ax.legend(title="Dự án", bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=9)
            plt.tight_layout(rect=[0, 0, 0.85, 1]) # Điều chỉnh bố cục cho chú giải
        elif mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
            total_hours_col_name = [col for col in df_plot.columns if 'Total Hours' in col][0] if [col for col in df_plot.columns if 'Total Hours' in col] else 'Total Hours'
            
            if 'MonthName' in df_plot.columns and len(comparison_config_inner['years']) == 1:
                df_plot.plot(kind='line', x='MonthName', y=total_hours_col_name, ax=ax, marker='o', color='green')
                ax.set_xticklabels(df_plot['MonthName'], rotation=45, ha="right", fontsize=10)
            elif 'Year' in df_plot.columns and not comparison_config_inner['months'] and len(comparison_config_inner['years']) > 1:
                df_plot.plot(kind='line', x='Year', y=total_hours_col_name, ax=ax, marker='o', color='blue')
                ax.set_xticklabels(df_plot['Year'], rotation=45, ha="right", fontsize=10)
            else:
                print("DEBUG: Không xác định được loại biểu đồ hợp lệ cho 'So sánh một dự án theo thời gian'.")
                plt.close(fig)
                return None
        else:
            print(f"DEBUG: Chế độ so sánh không xác định '{mode}'. Bỏ qua việc tạo biểu đồ.")
            plt.close(fig)
            return None

        ax.set_title(title, fontsize=12, pad=20)
        ax.set_xlabel(x_label, fontsize=10)
        ax.set_ylabel(y_label, fontsize=10)
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        plt.tight_layout()
        fig.savefig(img_path, dpi=150)
        plt.close(fig)
        return img_path

    try:
        config_info = {
            "Chế độ so sánh": comparison_mode,
            "Năm": ', '.join(map(str, comparison_config.get('years', []))) if comparison_config.get('years') else "N/A",
            "Tháng": ', '.join(comparison_config.get('months', [])) if comparison_config.get('months') else "Tất cả",
            "Dự án": ', '.join(comparison_config.get('selected_projects', [])) if comparison_config.get('selected_projects') else "N/A"
        }

        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'Liberation Sans']
        plt.rcParams['axes.unicode_minus'] = False

        chart_title = ""
        x_label = ""
        y_label = "Giờ"

        if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
            chart_title = f"Tổng giờ cho các dự án trong {comparison_config.get('months', [''])[0]}, {comparison_config.get('years', [''])[0]}"
            x_label = "Dự án"
        elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
            chart_title = f"Giờ hàng tháng cho các dự án trong {comparison_config.get('years', [''])[0]}"
            x_label = "Tháng"
        elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
            selected_proj_name = comparison_config.get('selected_projects', [''])[0]
            if len(comparison_config.get('years', [])) == 1 and comparison_config.get('months'):
                chart_title = f"Tổng giờ cho {selected_proj_name} trong {comparison_config.get('years', [''])[0]} theo tháng"
                x_label = "Tháng"
            elif len(comparison_config.get('years', [])) > 1 and not comparison_config.get('months'):
                chart_title = f"Tổng giờ cho {selected_proj_name} qua các năm"
                x_label = "Năm"
            else:
                print("DEBUG: Cấu hình không hợp lệ cho 'So sánh một dự án theo thời gian'.")
                return False
        else:
            print(f"DEBUG: Chế độ so sánh không xác định: {comparison_mode}")
            return False

        chart_img_path = os.path.join(tmp_dir, f"comparison_chart_{sanitize_filename(comparison_mode)}.png")
        generated_chart_path = create_comparison_chart(df_comparison, comparison_mode, chart_title, x_label, y_label, chart_img_path, comparison_config)
        
        if generated_chart_path:
            charts_for_pdf.append((generated_chart_path, chart_title, ""))

        if not charts_for_pdf:
            print("Cảnh báo: Không có biểu đồ nào được tạo để đưa vào PDF. PDF có thể trống.")
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, "BÁO CÁO GIỜ TRIAC - SO SÁNH", ln=True, align='C')
            pdf.set_font("helvetica", '', 12)
            pdf.cell(0, 10, f"Được tạo vào: {datetime.datetime.today().strftime('%Y-%m-%d')}", ln=True, align='C')
            pdf.ln(10)
            pdf.set_font("helvetica", '', 11)
            for key, value in config_info.items():
                pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')
            pdf.cell(0, 10, "Không có biểu đồ nào được tạo cho báo cáo này.", ln=True, align='C')
            pdf.output(pdf_file_path, "F")
            return True

        create_pdf_from_charts_comp(charts_for_pdf, pdf_file_path, "BÁO CÁO GIỜ TRIAC - SO SÁNH", config_info, logo_path)
        return True

    except Exception as e:
        print(f"Lỗi khi tạo báo cáo PDF so sánh: {e}")
        return False
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)
