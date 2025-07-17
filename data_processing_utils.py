import pandas as pd
import datetime
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from matplotlib import pyplot as plt
import tempfile
import re
import shutil
import streamlit as st # Import Streamlit

# Hàm hỗ trợ làm sạch tên file/sheet
def sanitize_filename(name):
    """
    Sanitizes a string to be safe for use as a filename and Excel sheet name.
    Removes invalid characters and limits length to 31 characters.
    """
    invalid_chars = re.compile(r'[\\/*?[\]:;|=,<>]')
    s = invalid_chars.sub("_", str(name))
    s = ''.join(c for c in s if c.isprintable()) # Loại bỏ các ký tự không in được
    return s[:31] # Giới hạn 31 ký tự cho tên sheet trong Excel

def setup_paths():
    """Thiết lập các đường dẫn file đầu ra. Không còn cần thiết cho file đầu vào nếu dùng st.file_uploader."""
    today = datetime.datetime.today().strftime('%Y%m%d')
    return {
        'output_file': f"Time_report_Standard_{today}.xlsx",
        'pdf_report': f"Time_report_Standard_{today}.pdf",
        'comparison_output_file': f"Time_report_Comparison_{today}.xlsx",
        'comparison_pdf_report': f"Time_report_Comparison_{today}.pdf",
        'logo_path': "triac_logo.png" # Đường dẫn logo, giả định nằm trong cùng thư mục
    }

@st.cache_data(show_spinner=False)
def read_configs(uploaded_file): # Tham số là đối tượng uploaded_file của Streamlit
    """Đọc cấu hình từ file Excel đã tải lên."""
    config_data = {}
    try:
        # Sử dụng BytesIO để đọc file trong bộ nhớ mà không cần lưu vào đĩa
        # Nếu uploaded_file là đường dẫn thì bỏ qua BytesIO
        if hasattr(uploaded_file, 'read'): # Kiểm tra nếu là BytesIO object (từ st.file_uploader)
            workbook = load_workbook(uploaded_file, read_only=True)
        else: # Nếu là đường dẫn file trên hệ thống
            workbook = load_workbook(uploaded_file, read_only=True)


        # --- Đọc Config_Year_Mode ---
        if 'Config_Year_Mode' in workbook.sheetnames:
            df_year_mode = pd.read_excel(uploaded_file, sheet_name='Config_Year_Mode', engine='openpyxl')
            
            if 'Key' in df_year_mode.columns and 'Value' in df_year_mode.columns:
                # Đảm bảo cột 'Key' là chuỗi để so sánh
                df_year_mode['Key'] = df_year_mode['Key'].astype(str).str.lower()
                config_year_mode_dict = df_year_mode.set_index('Key')['Value'].to_dict()
                
                config_data['mode'] = str(config_year_mode_dict.get('mode', 'year')).strip().lower()
                
                default_year_val = config_year_mode_dict.get('year', datetime.datetime.now().year)
                try:
                    # Đảm bảo 'Year' là số nguyên
                    config_data['year'] = int(default_year_val)
                except (ValueError, TypeError):
                    st.warning(f"Giá trị 'Year' trong cấu hình ('{default_year_val}') không hợp lệ. Sử dụng năm hiện tại.")
                    config_data['year'] = datetime.datetime.now().year

                months_val = config_year_mode_dict.get('months', '')
                if pd.isna(months_val) or months_val == '':
                    config_data['months'] = []
                else:
                    config_data['months'] = [m.strip().capitalize() for m in str(months_val).split(',')]
                
                config_data['weeks'] = config_year_mode_dict.get('weeks', None) # Giữ nguyên nếu có cột 'Weeks'

            else:
                st.error("Lỗi cấu hình: Sheet 'Config_Year_Mode' thiếu cột 'Key' hoặc 'Value'.")
                config_data['mode'] = 'year'
                config_data['year'] = datetime.datetime.now().year
                config_data['months'] = []
                config_data['weeks'] = None

        else:
            st.warning(f"Cảnh báo: Sheet cấu hình 'Config_Year_Mode' không tìm thấy trong file. Sử dụng các giá trị mặc định.")
            config_data['mode'] = 'year'
            config_data['year'] = datetime.datetime.now().year
            config_data['months'] = []
            config_data['weeks'] = None

        # --- Đọc Config_Project_Filter ---
        if 'Config_Project_Filter' in workbook.sheetnames:
            df_project_filter = pd.read_excel(uploaded_file, sheet_name='Config_Project_Filter', engine='openpyxl')
            if 'Project Name' in df_project_filter.columns:
                # Đảm bảo cột 'Project Name' và 'Include' là chuỗi
                df_project_filter['Project Name'] = df_project_filter['Project Name'].astype(str)
                if 'Include' in df_project_filter.columns:
                    df_project_filter['Include'] = df_project_filter['Include'].astype(str).str.lower()
                else:
                    st.warning("Cảnh báo: Sheet 'Config_Project_Filter' thiếu cột 'Include'. Sẽ không thể lọc dự án theo cột này.")
                    df_project_filter['Include'] = 'yes' # Mặc định bao gồm tất cả nếu không có cột 'Include'
                
                config_data['project_filter_df'] = df_project_filter
            else:
                st.warning("Cảnh báo: Sheet 'Config_Project_Filter' không có cột 'Project Name'. Không áp dụng bộ lọc dự án từ cấu hình.")
                config_data['project_filter_df'] = pd.DataFrame(columns=['Project Name', 'Include'])
        else:
            st.warning(f"Cảnh báo: Sheet cấu hình 'Config_Project_Filter' không tìm thấy trong file. Không áp dụng bộ lọc dự án từ cấu hình.")
            config_data['project_filter_df'] = pd.DataFrame(columns=['Project Name', 'Include'])

    except Exception as e:
        st.error(f"Lỗi khi đọc các sheet cấu hình: {e}. Vui lòng đảm bảo định dạng file Excel chính xác.")
        # Trả về cấu hình mặc định để ứng dụng không bị crash hoàn toàn
        config_data = {
            'mode': 'year',
            'year': datetime.datetime.now().year,
            'months': [],
            'weeks': None,
            'project_filter_df': pd.DataFrame(columns=['Project Name', 'Include'])
        }
    return config_data

@st.cache_data(show_spinner=False)
def load_raw_data(uploaded_file): # Tham số là đối tượng uploaded_file của Streamlit
    """Tải dữ liệu thô từ file Excel đã tải lên."""
    try:
        # Tương tự như read_configs, xử lý uploaded_file
        if hasattr(uploaded_file, 'read'):
            workbook = load_workbook(uploaded_file, read_only=True)
        else:
            workbook = load_workbook(uploaded_file, read_only=True)

        if 'Raw Data' not in workbook.sheetnames:
            st.error(f"Lỗi: Sheet 'Raw Data' không tìm thấy trong file Excel '{uploaded_file.name}'.")
            return pd.DataFrame()
        
        raw_df = pd.read_excel(uploaded_file, sheet_name='Raw Data', engine='openpyxl')
        
        # Làm sạch tên cột: loại bỏ khoảng trắng đầu/cuối, thay khoảng trắng bằng _, chuyển thành chữ thường
        # Chú ý: bạn đã rename 'Project Name' thành 'Project name' trong code trước, hãy thống nhất
        # raw_df.columns = raw_df.columns.str.strip().str.replace(' ', '_').str.lower()
        # Dựa vào cấu trúc ảnh, tên cột gốc có vẻ có khoảng trắng.
        # Hãy làm sạch theo cách bạn đã làm nhưng giữ nguyên chữ hoa nếu cần cho các bước sau
        raw_df.columns = [col.strip() for col in raw_df.columns] # Chỉ strip khoảng trắng
        
        # Đổi tên cột thủ công dựa trên ảnh chụp màn hình và code cũ của bạn
        # Cột 'Hou' và 'Team member' không có trong ảnh 'Raw Data' mới.
        # Có 'ID hours' và 'Employee name'
        # Dựa trên ảnh: 'Date', 'Team', 'Team leader', 'Team member', 'ID hours', 'Project name', 'Project code', 'Category', 'Workcentre', 'Task', 'Job'
        # Code cũ của bạn đổi: 'Hou': 'Hours', 'Team member': 'Employee', 'Project Name': 'Project name'
        # Cần điều chỉnh cho phù hợp với ảnh chụp màn hình mới nhất.
        
        # Đổi tên cột để khớp với tên bạn mong đợi trong code
        # Các cột từ ảnh chụp màn hình: 'Date', 'Month', 'Team leader', 'Team member', 'ID hours', 'Project name', 'Project code', 'Category', 'Workcentre', 'Task', 'Job'
        # Giữ nguyên tên gốc từ ảnh, chỉ chuyển đổi các tên cần thiết cho code.
        rename_map = {
            'ID hours': 'Hours', # Giả định 'ID hours' là 'Hours' bạn muốn tính
            'Team member': 'Employee name', # Để khớp với Employee name trong code Streamlit
            'Project name': 'Project name' # Giữ nguyên
        }
        # Chỉ đổi tên các cột có trong DataFrame
        raw_df.rename(columns={k: v for k, v in rename_map.items() if k in raw_df.columns}, inplace=True)

        # --- Xử lý cột 'Date' một cách mạnh mẽ hơn để tránh lỗi float-str ---
        if 'Date' in raw_df.columns:
            # Chuyển đổi tất cả các giá trị trong cột 'Date' sang kiểu string trước khi chuyển đổi sang datetime.
            # Điều này giúp tránh lỗi float-str nếu có giá trị số không phải ngày hoặc các kiểu hỗn hợp khác.
            raw_df['Date'] = raw_df['Date'].astype(str)
            raw_df['Date'] = pd.to_datetime(raw_df['Date'], errors='coerce', dayfirst=True) # Thử dayfirst=True nếu định dạng là DD/MM/YYYY
            
            invalid_dates_count = raw_df['Date'].isna().sum()
            if invalid_dates_count > 0:
                st.warning(f"Có {invalid_dates_count} giá trị không hợp lệ trong cột 'Date' của sheet 'Raw Data'. Các hàng này sẽ bị loại bỏ.")
                # st.write("Một số giá trị 'Date' không hợp lệ gốc:", raw_df[raw_df['Date'].isna()]['Original_Date_Column'].head()) # Nếu bạn giữ bản sao
            
            raw_df = raw_df.dropna(subset=['Date']) # Loại bỏ các hàng có ngày không hợp lệ

            if raw_df.empty:
                st.warning("Sau khi loại bỏ các hàng có ngày không hợp lệ, DataFrame 'Raw Data' trở nên rỗng.")
                return pd.DataFrame()
            
            raw_df['Year'] = raw_df['Date'].dt.year
            raw_df['MonthName'] = raw_df['Date'].dt.strftime('%B') # Tên tháng đầy đủ (January, February...)
            raw_df['Week'] = raw_df['Date'].dt.isocalendar().week.astype(int)
        else:
            st.error("Lỗi: Cột 'Date' không tìm thấy trong sheet 'Raw Data'. Vui lòng kiểm tra file Excel của bạn.")
            return pd.DataFrame()

        # --- Đảm bảo cột 'Hours' là số và xử lý các kiểu hỗn hợp ---
        if 'Hours' in raw_df.columns:
            # Chuyển đổi cột 'Hours' sang kiểu số, ép buộc lỗi thành NaN, sau đó điền 0 cho NaN
            raw_df['Hours'] = pd.to_numeric(raw_df['Hours'], errors='coerce').fillna(0)
        else:
            st.warning("Cảnh báo: Cột 'ID hours' (hoặc 'Hours') không tìm thấy trong sheet 'Raw Data'. Đặt cột 'Hours' về 0.")
            raw_df['Hours'] = 0 # Thêm cột nếu nó bị thiếu

        # Kiểm tra sự tồn tại của các cột cần thiết khác sau khi đổi tên
        required_cols = ['Project name', 'Employee name', 'Workcentre', 'Task', 'Hours', 'Year', 'MonthName', 'Week']
        for col in required_cols:
            if col not in raw_df.columns:
                st.warning(f"Cảnh báo: Cột '{col}' không tìm thấy trong dữ liệu thô. Có thể ảnh hưởng đến báo cáo.")
                # Thêm cột rỗng nếu thiếu để tránh lỗi Key Error sau này
                raw_df[col] = '' if col not in ['Hours', 'Year', 'Week'] else 0

        return raw_df
    except Exception as e:
        st.error(f"Lỗi khi tải dữ liệu thô từ file Excel: {e}. Vui lòng kiểm tra định dạng và nội dung file Excel.")
        return pd.DataFrame()

def apply_filters(df, config):
    """Áp dụng các bộ lọc dữ liệu dựa trên cấu hình."""
    df_filtered = df.copy()

    # Chuyển đổi năm thành kiểu số nguyên để so sánh an toàn
    if 'Year' in df_filtered.columns:
        df_filtered['Year'] = pd.to_numeric(df_filtered['Year'], errors='coerce').fillna(-1).astype(int)
    else:
        st.warning("Cột 'Year' không tồn tại trong DataFrame để áp dụng bộ lọc.")
        return pd.DataFrame() # Hoặc xử lý khác tùy ý

    if 'mode' in config and config['mode'] == 'comparison_years' and 'years' in config and config['years']: # Dành cho so sánh nhiều năm
        # Đảm bảo config['years'] là danh sách các số nguyên
        filter_years = [int(y) for y in config['years'] if pd.notna(y) and str(y).replace('.', '').isdigit()]
        df_filtered = df_filtered[df_filtered['Year'].isin(filter_years)]
    elif 'year' in config and config['year']: # Dành cho báo cáo tiêu chuẩn một năm
        df_filtered = df_filtered[df_filtered['Year'] == config['year']]

    if 'months' in config and config['months']:
        # Đảm bảo cột 'MonthName' là chuỗi để so sánh
        if 'MonthName' in df_filtered.columns:
            df_filtered['MonthName'] = df_filtered['MonthName'].astype(str)
            df_filtered = df_filtered[df_filtered['MonthName'].isin(config['months'])]
        else:
            st.warning("Cột 'MonthName' không tồn tại trong DataFrame để áp dụng bộ lọc tháng.")


    if not config['project_filter_df'].empty:
        # Lấy danh sách các dự án được chọn để bao gồm (Include = 'yes')
        included_projects_df = config['project_filter_df'][
            (config['project_filter_df']['Include'].astype(str).str.lower() == 'yes') | 
            (config['project_filter_df']['Include'].astype(str).str.lower() == 'true')
        ]
        if not included_projects_df.empty:
            selected_project_names = included_projects_df['Project Name'].tolist()
            if 'Project name' in df_filtered.columns:
                df_filtered['Project name'] = df_filtered['Project name'].astype(str) # Đảm bảo là chuỗi
                df_filtered = df_filtered[df_filtered['Project name'].isin(selected_project_names)]
            else:
                st.warning("Cột 'Project name' không tồn tại trong DataFrame để áp dụng bộ lọc dự án.")
                return pd.DataFrame()
        else:
            st.info("Không có dự án nào được chọn để đưa vào từ cấu hình. Trả về DataFrame rỗng.")
            return pd.DataFrame(columns=df.columns) # Trả về DataFrame rỗng nếu không có dự án nào được chọn
    else:
        st.info("Không có cấu hình lọc dự án. Trả về DataFrame rỗng.")
        return pd.DataFrame(columns=df.columns)

    if df_filtered.empty:
        st.info("Không có dữ liệu nào khớp với các bộ lọc đã chọn.")

    return df_filtered

def export_report(df, config, output_file_path):
    """Xuất báo cáo tiêu chuẩn ra file Excel."""
    mode = config.get('mode', 'year')
    
    groupby_cols = []
    if mode == 'year':
        groupby_cols = ['Year', 'Project name']
    elif mode == 'month':
        groupby_cols = ['Year', 'MonthName', 'Project name']
    else: # week mode
        groupby_cols = ['Year', 'Week', 'Project name']

    # Kiểm tra sự tồn tại của các cột cần thiết trước khi groupby
    for col in groupby_cols + ['Hours']:
        if col not in df.columns:
            st.error(f"Lỗi: Cột '{col}' không tồn tại trong DataFrame. Không thể tạo báo cáo.")
            return False

    if df.empty:
        st.warning("Cảnh báo: DataFrame đã lọc trống, không có báo cáo nào được tạo.")
        return False

    summary = df.groupby(groupby_cols)['Hours'].sum().reset_index()

    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            summary.to_excel(writer, sheet_name='Summary', index=False)

            wb = writer.book
            ws = wb['Summary']
            
            if len(summary) > 0:
                data_col_idx = summary.columns.get_loc('Hours') + 1
                cats_col_idx = summary.columns.get_loc('Project name') + 1

                data_ref = Reference(ws, min_col=data_col_idx, min_row=2, max_row=ws.max_row)
                cats_ref = Reference(ws, min_col=cats_col_idx, min_row=2, max_row=ws.max_row)

                chart = BarChart()
                chart.title = f"Tổng số giờ theo dự án ({mode.capitalize()})"
                chart.x_axis.title = "Dự án"
                chart.y_axis.title = "Giờ"
                
                chart.add_data(data_ref, titles_from_data=False) 
                chart.set_categories(cats_ref)
                ws.add_chart(chart, "F2")

            # Tạo các sheet chi tiết cho từng dự án
            for project in df['Project name'].unique():
                df_proj = df[df['Project name'] == project].copy() # Tạo bản sao để tránh SettingWithCopyWarning
                sheet_title = sanitize_filename(project)
                
                # Kiểm tra nếu sheet đã tồn tại, tránh ghi đè hoặc tạo sheet trùng tên
                if sheet_title in wb.sheetnames:
                    ws_proj = wb[sheet_title]
                else:
                    ws_proj = wb.create_sheet(title=sheet_title)

                # Summary theo Task
                summary_task = df_proj.groupby('Task')['Hours'].sum().reset_index().sort_values('Hours', ascending=False)
                
                if not summary_task.empty:
                    ws_proj.append(['Task', 'Hours']) # Thêm header cho summary task
                    for row_data in dataframe_to_rows(summary_task, index=False, header=False):
                        ws_proj.append(row_data)

                    chart_task = BarChart()
                    chart_task.title = f"{project} - Giờ theo Task"
                    chart_task.x_axis.title = "Task"
                    chart_task.y_axis.title = "Giờ"
                    task_len = len(summary_task)
                    
                    data_ref_task = Reference(ws_proj, min_col=2, min_row=1, max_row=task_len + 1)
                    cats_ref_task = Reference(ws_proj, min_col=1, min_row=2, max_row=task_len + 1)
                    chart_task.add_data(data_ref_task, titles_from_data=True)
                    chart_task.set_categories(cats_ref_task)
                    ws_proj.add_chart(chart_task, f"E1")

                # Dữ liệu thô của dự án
                start_row_raw_data = ws_proj.max_row + 2
                if not summary_task.empty: # Nếu có biểu đồ task, thêm khoảng cách
                    start_row_raw_data += 15 # Đảm bảo đủ chỗ cho biểu đồ và thông tin tóm tắt task

                # Ghi header trước khi ghi dữ liệu
                for c_idx, col_name in enumerate(df_proj.columns):
                    ws_proj.cell(row=start_row_raw_data, column=c_idx + 1, value=col_name)

                for r_idx, r in enumerate(dataframe_to_rows(df_proj, index=False, header=False)):
                    for c_idx, cell_val in enumerate(r):
                        ws_proj.cell(row=start_row_raw_data + r_idx + 1, column=c_idx + 1, value=cell_val)
            
            # Tạo sheet Config_Info
            ws_config = wb.create_sheet("Config_Info")
            ws_config['A1'], ws_config['B1'] = "Chế độ", config.get('mode', 'N/A').capitalize()
            
            years_display = ', '.join(map(str, config.get('years', []))) if config.get('years') else str(config.get('year', 'N/A'))
            ws_config['A2'], ws_config['B2'] = "Năm", years_display
            
            months_display = ', '.join(config.get('months', [])) if config.get('months') else "Tất cả"
            ws_config['A3'], ws_config['B3'] = "Tháng", months_display
            
            if 'project_filter_df' in config and not config['project_filter_df'].empty:
                selected_projects_display = config['project_filter_df'][
                    (config['project_filter_df']['Include'].astype(str).str.lower() == 'yes') |
                    (config['project_filter_df']['Include'].astype(str).str.lower() == 'true')
                ]['Project Name'].tolist()
                ws_config['A4'], ws_config['B4'] = "Dự án được bao gồm", ', '.join(selected_projects_display)
            else:
                ws_config['A4'], ws_config['B4'] = "Dự án được bao gồm", "Không có dự án nào được chọn hoặc tìm thấy"

            # Xóa các sheet template
            for sheet_name in ['Raw Data', 'Config_Year_Mode', 'Config_Project_Filter']:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

            wb.save(output_file_path)
            st.success(f"Báo cáo Excel tiêu chuẩn đã được tạo thành công tại: {output_file_path}")
            return True
    except Exception as e:
        st.error(f"Lỗi khi xuất báo cáo tiêu chuẩn ra Excel: {e}.")
        return False

def export_pdf_report(df, config, pdf_report_path, logo_path):
    """Xuất báo cáo PDF tiêu chuẩn với các biểu đồ."""
    today_str = datetime.datetime.today().strftime("%Y-%m-%d")
    tmp_dir = tempfile.mkdtemp()
    charts_for_pdf = []

    def create_pdf_from_charts(charts_data, output_path, title, config_info, logo_path_inner):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font('helvetica', 'B', 16) 
        
        pdf.add_page()
        if os.path.exists(logo_path_inner):
            pdf.image(logo_path_inner, x=10, y=10, w=30)
        pdf.ln(40)
        pdf.cell(0, 10, title, ln=True, align='C')
        pdf.set_font("helvetica", '', 12) 
        pdf.ln(5)
        pdf.cell(0, 10, f"Generated on: {today_str}", ln=True, align='C')
        pdf.ln(10)
        pdf.set_font("helvetica", '', 11) 
        for key, value in config_info.items():
            pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')

        for img_path, chart_title, page_project_name in charts_data:
            if img_path and os.path.exists(img_path):
                pdf.add_page()
                if os.path.exists(logo_path_inner):
                    pdf.image(logo_path_inner, x=10, y=8, w=25)
                pdf.set_font("helvetica", 'B', 11) 
                pdf.set_y(35)
                if page_project_name:
                    pdf.cell(0, 10, f"Project: {page_project_name}", ln=True, align='C')
                pdf.cell(0, 10, chart_title, ln=True, align='C')
                pdf.image(img_path, x=10, y=45, w=190)

        pdf.output(output_path, "F")
        st.success(f"Báo cáo PDF đã được tạo thành công tại: {output_path}")

    try:
        projects = df['Project name'].unique() 

        config_info = {
            "Chế độ": config.get('mode', 'N/A').capitalize(),
            "Năm": ', '.join(map(str, config.get('years', []))) if config.get('years') else str(config.get('year', 'N/A')),
            "Tháng": ', '.join(config.get('months', [])) if config.get('months') else "Tất cả",
            "Dự án được bao gồm": ', '.join(config['project_filter_df']['Project Name'].tolist()) if 'project_filter_df' in config and not config['project_filter_df'].empty else "Không có dự án nào được chọn hoặc tìm thấy"
        }

        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'Liberation Sans']
        plt.rcParams['axes.unicode_minus'] = False 

        for project in projects:
            safe_project = sanitize_filename(project)
            df_proj = df[df['Project name'] == project].copy() # Tạo bản sao

            if 'Workcentre' in df_proj.columns and not df_proj['Workcentre'].empty:
                workcentre_summary = df_proj.groupby('Workcentre')['Hours'].sum().sort_values(ascending=False)
                if not workcentre_summary.empty and workcentre_summary.sum() > 0:
                    fig, ax = plt.subplots(figsize=(10, 5))
                    workcentre_summary.plot(kind='barh', color='skyblue', ax=ax)
                    ax.set_title(f"{project} - Giờ theo Workcentre", fontsize=9)
                    ax.tick_params(axis='y', labelsize=8)
                    ax.set_xlabel("Giờ")
                    ax.set_ylabel("Workcentre")
                    wc_img_path = os.path.join(tmp_dir, f"{safe_project}_wc.png")
                    plt.tight_layout()
                    fig.savefig(wc_img_path, dpi=150)
                    plt.close(fig)
                    charts_for_pdf.append((wc_img_path, f"{project} - Giờ theo Workcentre", project))

            if 'Task' in df_proj.columns and not df_proj['Task'].empty:
                task_summary = df_proj.groupby('Task')['Hours'].sum().sort_values(ascending=False)
                if not task_summary.empty and task_summary.sum() > 0:
                    fig, ax = plt.subplots(figsize=(10, 6))
                    task_summary.plot(kind='barh', color='lightgreen', ax=ax)
                    ax.set_title(f"{project} - Giờ theo Task", fontsize=9)
                    ax.tick_params(axis='y', labelsize=8)
                    ax.set_xlabel("Giờ")
                    ax.set_ylabel("Task")
                    task_img_path = os.path.join(tmp_dir, f"{safe_project}_task.png")
                    plt.tight_layout()
                    fig.savefig(task_img_path, dpi=150)
                    plt.close(fig)
                    charts_for_pdf.append((task_img_path, f"{project} - Giờ theo Task", project))

        if not charts_for_pdf:
            st.warning("Cảnh báo: Không có biểu đồ nào được tạo để đưa vào PDF. PDF có thể trống.")
            # Vẫn tạo PDF trống với thông tin config
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, "TRIAC TIME REPORT - STANDARD", ln=True, align='C')
            pdf.set_font("helvetica", '', 12)
            pdf.cell(0, 10, f"Generated on: {today_str}", ln=True, align='C')
            pdf.ln(10)
            pdf.set_font("helvetica", '', 11)
            for key, value in config_info.items():
                pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')
            pdf.cell(0, 10, "Không có biểu đồ nào được tạo cho báo cáo này.", ln=True, align='C')
            pdf.output(pdf_report_path, "F")
            return True
            
        create_pdf_from_charts(charts_for_pdf, pdf_report_path, "TRIAC TIME REPORT - STANDARD", config_info, logo_path)
        return True
    except Exception as e:
        st.error(f"Lỗi khi tạo báo cáo PDF: {e}.")
        return False
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir) # Dọn dẹp thư mục tạm thời

def apply_comparison_filters(df_raw, comparison_config, comparison_mode):
    """Áp dụng bộ lọc và tạo DataFrame tóm tắt cho báo cáo so sánh."""
    # Lấy các bộ lọc từ comparison_config, với kiểm tra an toàn
    years = comparison_config.get('years', [])
    months = comparison_config.get('months', [])
    selected_projects = comparison_config.get('selected_projects', [])

    df_filtered = df_raw.copy()

    # Kiểm tra và chuyển đổi kiểu dữ liệu của cột 'Year' và 'MonthName'
    if 'Year' in df_filtered.columns:
        df_filtered['Year'] = pd.to_numeric(df_filtered['Year'], errors='coerce').fillna(-1).astype(int)
    else:
        st.warning("Cột 'Year' không tồn tại trong DataFrame để áp dụng bộ lọc so sánh.")
        return pd.DataFrame(), "Cột 'Year' không tồn tại."

    if 'MonthName' in df_filtered.columns:
        df_filtered['MonthName'] = df_filtered['MonthName'].astype(str)
    else:
        st.warning("Cột 'MonthName' không tồn tại trong DataFrame để áp dụng bộ lọc so sánh.")
        return pd.DataFrame(), "Cột 'MonthName' không tồn tại."
    
    if 'Project name' in df_filtered.columns:
        df_filtered['Project name'] = df_filtered['Project name'].astype(str)
    else:
        st.warning("Cột 'Project name' không tồn tại trong DataFrame để áp dụng bộ lọc so sánh.")
        return pd.DataFrame(), "Cột 'Project name' không tồn tại."


    if years:
        df_filtered = df_filtered[df_filtered['Year'].isin(years)]
    
    if months:
        df_filtered = df_filtered[df_filtered['MonthName'].isin(months)]
    
    if selected_projects:
        df_filtered = df_filtered[df_filtered['Project name'].isin(selected_projects)]
    else: 
        st.warning("Vui lòng chọn ít nhất một dự án để so sánh.")
        return pd.DataFrame(), "Vui lòng chọn ít nhất một dự án để so sánh."

    if df_filtered.empty:
        return pd.DataFrame(), f"Không tìm thấy dữ liệu cho chế độ so sánh: {comparison_mode} với các lựa chọn hiện tại."

    title = ""

    if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
        if len(years) != 1 or len(months) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), "Vui lòng chọn MỘT năm, MỘT tháng và ít nhất HAI dự án cho chế độ này."
        
        df_comparison = df_filtered.groupby('Project name')['Hours'].sum().reset_index()
        df_comparison.rename(columns={'Hours': 'Total Hours'}, inplace=True)
        title = f"So sánh giờ giữa các dự án trong {months[0]}, năm {years[0]}"
        return df_comparison, title

    elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
        if len(years) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), "Vui lòng chọn MỘT năm và ít nhất HAI dự án cho chế độ này."
        
        df_comparison = df_filtered.groupby(['Project name', 'MonthName'])['Hours'].sum().unstack(fill_value=0)
        
        month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        existing_months = [m for m in month_order if m in df_comparison.columns]
        df_comparison = df_comparison[existing_months]

        df_comparison = df_comparison.reset_index().rename(columns={'index': 'Project Name'})
        
        df_comparison['Total Hours'] = df_comparison[existing_months].sum(axis=1)

        # Thêm hàng tổng nếu cần cho báo cáo Excel
        df_comparison.loc['Total'] = df_comparison[existing_months + ['Total Hours']].sum()
        df_comparison.loc['Total', 'Project Name'] = 'Total'

        title = f"So sánh giờ giữa các dự án trong năm {years[0]} (theo tháng)"
        return df_comparison, title

    elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
        # Đã xác thực rằng selected_projects chỉ có 1 trong main_optimized.py
        if len(selected_projects) != 1:
            return pd.DataFrame(), "Lỗi: Internal - Vui lòng chọn CHỈ MỘT dự án cho chế độ này."

        selected_project_name = selected_projects[0]

        if len(years) == 1 and len(months) > 0:
            df_comparison = df_filtered.groupby('MonthName')['Hours'].sum().reset_index()
            df_comparison.rename(columns={'Hours': f'Total Hours for {selected_project_name}'}, inplace=True)
            
            month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            df_comparison['MonthName'] = pd.Categorical(df_comparison['MonthName'], categories=month_order, ordered=True)
            df_comparison = df_comparison.sort_values('MonthName').reset_index(drop=True)
            
            df_comparison['Project Name'] = selected_project_name # Thêm cột này để nhất quán
            title = f"Tổng giờ dự án {selected_project_name} qua các tháng trong năm {years[0]}"
            return df_comparison, title

        elif len(years) > 1 and not months:
            df_comparison = df_filtered.groupby('Year')['Hours'].sum().reset_index()
            df_comparison.rename(columns={'Hours': f'Total Hours for {selected_project_name}'}, inplace=True)
            df_comparison['Year'] = df_comparison['Year'].astype(str) # Chuyển năm thành chuỗi cho trục X
            
            df_comparison['Project Name'] = selected_project_name # Thêm cột này để nhất quán
            title = f"Tổng giờ dự án {selected_project_name} qua các năm"
            return df_comparison, title

        else:
            return pd.DataFrame(), "Cấu hình so sánh dự án qua thời gian không hợp lệ. Vui lòng chọn một năm với nhiều tháng, HOẶC nhiều năm."
            
    return pd.DataFrame(), "Chế độ so sánh không hợp lệ."

def export_comparison_report(df_comparison, comparison_config, output_file_path, comparison_mode):
    """Xuất báo cáo so sánh ra file Excel."""
    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            if df_comparison.empty:
                empty_df_for_excel = pd.DataFrame({"Message": ["Không có dữ liệu để hiển thị với các bộ lọc đã chọn."]})
                empty_df_for_excel.to_excel(writer, sheet_name='Comparison Report', index=False)
            else:
                df_comparison.to_excel(writer, sheet_name='Comparison Report', index=False) 

            wb = writer.book
            ws = wb['Comparison Report']

            # Thêm thông tin cấu hình vào báo cáo Excel
            data_last_row = ws.max_row
            info_row = data_last_row + 2 

            ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=4)
            ws.cell(row=info_row, column=1, value=f"BÁO CÁO SO SÁNH: {comparison_mode}").font = ws.cell(row=info_row, column=1).font.copy(bold=True, size=14)
            info_row += 1

            ws.cell(row=info_row, column=1, value="Năm:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(map(str, comparison_config.get('years', []))))
            info_row += 1
            ws.cell(row=info_row, column=1, value="Tháng:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('months', [])))
            info_row += 1
            ws.cell(row=info_row, column=1, value="Dự án:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('selected_projects', [])))

            if not df_comparison.empty and len(df_comparison) > 0:
                chart = None
                data_start_row = 2 
                
                df_chart_data = df_comparison.copy() 
                # Loại bỏ hàng 'Total' để không vẽ vào biểu đồ
                if 'Project Name' in df_chart_data.columns and 'Total' in df_chart_data['Project Name'].values:
                    df_chart_data = df_chart_data[df_chart_data['Project Name'] != 'Total']
                elif 'Year' in df_chart_data.columns and 'Total' in df_chart_data['Year'].values:
                    df_chart_data = df_chart_data[df_chart_data['Year'] != 'Total']
                
                if df_chart_data.empty: 
                    st.warning("Không có đủ dữ liệu để vẽ biểu đồ so sánh sau khi loại bỏ hàng tổng.")
                    wb.save(output_file_path)
                    return True

                max_row_chart = data_start_row + len(df_chart_data) - 1

                if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
                    chart = BarChart()
                    chart.title = "So sánh giờ theo dự án"
                    chart.x_axis.title = "Dự án"
                    chart.y_axis.title = "Giờ"
                    
                    data_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Total Hours') + 1, min_row=data_start_row, max_row=max_row_chart)
                    cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Project Name') + 1, min_row=data_start_row, max_row=max_row_chart) 
                    
                    chart.add_data(data_ref, titles_from_data=False) 
                    chart.set_categories(cats_ref)
                
                elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
                    chart = LineChart()
                    chart.title = "So sánh giờ theo dự án và tháng"
                    chart.x_axis.title = "Tháng"
                    chart.y_axis.title = "Giờ"

                    month_cols = [col for col in df_comparison.columns if col not in ['Project Name', 'Total Hours']]
                    
                    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
                    ordered_month_cols = [m for m in month_order if m in month_cols]

                    if ordered_month_cols:
                        min_col_month_index = df_comparison.columns.get_loc(ordered_month_cols[0])
                        max_col_month_index = df_comparison.columns.get_loc(ordered_month_cols[-1])
                        min_col_month = min_col_month_index + 1 
                        max_col_month = max_col_month_index + 1
                        cats_ref = Reference(ws, min_col=min_col_month, min_row=1, max_col=max_col_month)
                    else:
                        st.warning("Không tìm thấy cột tháng để tạo biểu đồ.")
                        wb.save(output_file_path)
                        return True
                    
                    for r_idx, project_name in enumerate(df_chart_data['Project Name']):
                        series_ref = Reference(ws, min_col=min_col_month, 
                                                min_row=data_start_row + r_idx, 
                                                max_col=max_col_month, 
                                                max_row=data_start_row + r_idx)
                        title_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Project Name') + 1, 
                                            min_row=data_start_row + r_idx, 
                                            max_row=data_start_row + r_idx)
                        chart.add_data(series_ref, titles_from_data=True)
                        chart.series[r_idx].title = title_ref
                    
                    chart.set_categories(cats_ref)

                elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
                    total_hours_col_name = [col for col in df_comparison.columns if 'Total Hours' in col][0] if [col for col in df_comparison.columns if 'Total Hours' in col] else 'Total Hours'
                    
                    if 'MonthName' in df_comparison.columns and len(comparison_config.get('years', [])) == 1:
                        chart = BarChart() # BarChart phù hợp cho tổng giờ theo từng tháng cụ thể
                        chart.title = f"Tổng giờ dự án {comparison_config['selected_projects'][0]} năm {comparison_config['years'][0]} theo tháng"
                        chart.x_axis.title = "Tháng"
                        chart.y_axis.title = "Giờ"
                        
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('MonthName') + 1, min_row=data_start_row, max_row=max_row_chart)
                        
                        chart.add_data(data_ref, titles_from_data=False) 
                        chart.set_categories(cats_ref)
                    elif 'Year' in df_comparison.columns and not comparison_config.get('months', []) and len(comparison_config.get('years', [])) > 1:
                        chart = LineChart() # LineChart phù hợp hơn cho xu hướng qua các năm
                        chart.title = f"Tổng giờ dự án {comparison_config['selected_projects'][0]} qua các năm"
                        chart.x_axis.title = "Năm"
                        chart.y_axis.title = "Giờ"
                        
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Year') + 1, min_row=data_start_row, max_row=max_row_chart)
                        
                        chart.add_data(data_ref, titles_from_data=False) 
                        chart.set_categories(cats_ref)
                    else:
                        raise ValueError("Không tìm thấy kích thước thời gian hợp lệ cho các danh mục biểu đồ trong chế độ so sánh qua tháng/năm.")

                if chart: 
                    chart_placement_row = info_row + 2
                    ws.add_chart(chart, f"A{chart_placement_row}")

            wb.save(output_file_path)
            st.success(f"Báo cáo Excel so sánh đã được tạo thành công tại: {output_file_path}")
            return True
    except Exception as e:
        st.error(f"Lỗi khi xuất báo cáo so sánh ra Excel: {e}.")
        return False

def export_comparison_pdf_report(df_comparison, comparison_config, pdf_file_path, comparison_mode, logo_path):
    """Xuất báo cáo PDF so sánh với biểu đồ."""
    tmp_dir = tempfile.mkdtemp()
    charts_for_pdf = []

    def create_pdf_from_charts_comp(charts_data, output_path, title, config_info, logo_path_inner):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font('helvetica', 'B', 16) 

        pdf.add_page()
        if os.path.exists(logo_path_inner):
            pdf.image(logo_path_inner, x=10, y=10, w=30)
        pdf.ln(40)
        pdf.cell(0, 10, title, ln=True, align='C')
        pdf.set_font("helvetica", '', 12) 
        pdf.ln(5)
        pdf.cell(0, 10, f"Generated on: {datetime.datetime.today().strftime('%Y-%m-%d')}", ln=True, align='C')
        pdf.ln(10)
        pdf.set_font("helvetica", '', 11) 
        for key, value in config_info.items():
            pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')

        for img_path, chart_title, page_project_name in charts_data:
            if img_path and os.path.exists(img_path):
                pdf.add_page()
                if os.path.exists(logo_path_inner):
                    pdf.image(logo_path_inner, x=10, y=8, w=25)
                pdf.set_font("helvetica", 'B', 11) 
                pdf.set_y(35)
                if page_project_name:
                    pdf.cell(0, 10, f"Project: {page_project_name}", ln=True, align='C')
                pdf.cell(0, 10, chart_title, ln=True, align='C')
                pdf.image(img_path, x=10, y=45, w=190)

        pdf.output(output_path, "F")
        st.success(f"Báo cáo PDF đã được tạo thành công tại: {output_path}")

    def create_comparison_chart(df, mode, title, x_label, y_label, img_path, comparison_config_inner):
        fig, ax = plt.subplots(figsize=(12, 7))  
        
        df_plot = df.copy()  
        
        # Loại bỏ hàng 'Total' nếu có để không ảnh hưởng đến biểu đồ
        if 'Project Name' in df_plot.columns and 'Total' in df_plot['Project Name'].values:
            df_plot = df_plot[df_plot['Project Name'] != 'Total']
        elif 'Year' in df_plot.columns and 'Total' in df_plot['Year'].values:
            df_plot = df_plot[df_plot['Year'] != 'Total']
        
        if df_plot.empty:
            st.warning(f"Không có dữ liệu để vẽ biểu đồ cho chế độ '{mode}' sau khi loại bỏ hàng tổng.")
            plt.close(fig)  
            return None 

        ax.set_ylim(bottom=0)
        
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'Liberation Sans']
        plt.rcParams['axes.unicode_minus'] = False 

        if mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
            # Biểu đồ cột so sánh các dự án trong một tháng
            df_plot.plot(kind='bar', x='Project Name', y='Total Hours', ax=ax, color='teal')
            ax.set_title(title, fontsize=12)
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            plt.xticks(rotation=45, ha='right')
            
        elif mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
            # Biểu đồ đường so sánh các dự án qua các tháng trong một năm
            month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            # Chuyển cột tháng sang kiểu Categorical để sắp xếp đúng thứ tự
            df_plot['MonthName'] = pd.Categorical(df_plot['MonthName'], categories=month_order, ordered=True)
            df_plot = df_plot.sort_values('MonthName')

            # Vẽ từng đường cho mỗi dự án
            for project in df_plot['Project Name'].unique():
                project_data = df_plot[df_plot['Project Name'] == project]
                ax.plot(project_data['MonthName'], project_data.drop(columns=['Project Name', 'Total Hours']).iloc[0], marker='o', label=project) # Lấy dữ liệu giờ của các tháng
            
            ax.set_title(title, fontsize=12)
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            ax.legend(title="Project")
            plt.xticks(rotation=45, ha='right')
            
        elif mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
            if 'MonthName' in df_plot.columns and len(comparison_config_inner.get('years', [])) == 1:
                # Biểu đồ cột/đường cho Tổng giờ theo Tháng (trong một năm)
                df_plot['MonthName'] = pd.Categorical(df_plot['MonthName'], categories=month_order, ordered=True)
                df_plot = df_plot.sort_values('MonthName')
                df_plot.plot(kind='bar', x='MonthName', y=df_plot.columns[1], ax=ax, color='darkorange') # Lấy tên cột giờ thứ 2
                ax.set_title(title, fontsize=12)
                ax.set_xlabel(x_label)
                ax.set_ylabel(y_label)
                plt.xticks(rotation=45, ha='right')
            elif 'Year' in df_plot.columns and not comparison_config_inner.get('months', []) and len(comparison_config_inner.get('years', [])) > 1:
                # Biểu đồ đường cho Tổng giờ theo Năm (qua nhiều năm)
                df_plot = df_plot.sort_values('Year')
                df_plot.plot(kind='line', x='Year', y=df_plot.columns[1], ax=ax, marker='o', color='purple') # Lấy tên cột giờ thứ 2
                ax.set_title(title, fontsize=12)
                ax.set_xlabel(x_label)
                ax.set_ylabel(y_label)
                plt.xticks(rotation=45, ha='right')
            else:
                st.error("Không thể tạo biểu đồ cho chế độ so sánh qua thời gian với cấu hình này.")
                plt.close(fig)
                return None
        else:
            st.warning("Chế độ so sánh không được hỗ trợ để tạo biểu đồ.")
            plt.close(fig)
            return None

        plt.tight_layout()
        fig.savefig(img_path, dpi=150)
        plt.close(fig)
        return img_path

    try:
        config_info = {
            "Chế độ": comparison_mode,
            "Năm": ', '.join(map(str, comparison_config.get('years', []))),
            "Tháng": ', '.join(comparison_config.get('months', [])) if comparison_config.get('months') else "Tất cả",
            "Dự án": ', '.join(comparison_config.get('selected_projects', []))
        }

        # Tạo biểu đồ chính cho báo cáo so sánh
        chart_title = ""
        x_label = ""
        y_label = "Giờ"
        img_filename = "comparison_chart.png"
        
        if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
            chart_title = f"Tổng giờ theo dự án trong {comparison_config.get('months', ['N/A'])[0]}, {comparison_config.get('years', ['N/A'])[0]}"
            x_label = "Dự án"
        elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
            chart_title = f"Tổng giờ theo dự án và tháng trong năm {comparison_config.get('years', ['N/A'])[0]}"
            x_label = "Tháng"
        elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
            selected_project_name = comparison_config.get('selected_projects', ['N/A'])[0]
            if len(comparison_config.get('years', [])) == 1 and len(comparison_config.get('months', [])) > 0:
                chart_title = f"Tổng giờ dự án {selected_project_name} qua các tháng trong năm {comparison_config.get('years', ['N/A'])[0]}"
                x_label = "Tháng"
            elif len(comparison_config.get('years', [])) > 1 and not comparison_config.get('months', []):
                chart_title = f"Tổng giờ dự án {selected_project_name} qua các năm"
                x_label = "Năm"
            else:
                st.error("Cấu hình không hợp lệ cho biểu đồ so sánh qua thời gian.")
                return False

        comparison_chart_path = create_comparison_chart(df_comparison, comparison_mode, chart_title, x_label, y_label, os.path.join(tmp_dir, img_filename), comparison_config)
        if comparison_chart_path:
            charts_for_pdf.append((comparison_chart_path, chart_title, ""))

        if not charts_for_pdf:
            st.warning("Cảnh báo: Không có biểu đồ nào được tạo để đưa vào PDF báo cáo so sánh. PDF có thể trống.")
            # Vẫn tạo PDF trống với thông tin config
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, "TRIAC TIME REPORT - COMPARISON", ln=True, align='C')
            pdf.set_font("helvetica", '', 12)
            pdf.cell(0, 10, f"Generated on: {datetime.datetime.today().strftime('%Y-%m-%d')}", ln=True, align='C')
            pdf.ln(10)
            pdf.set_font("helvetica", '', 11)
            for key, value in config_info.items():
                pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')
            pdf.cell(0, 10, "Không có biểu đồ nào được tạo cho báo cáo so sánh này.", ln=True, align='C')
            pdf.output(pdf_file_path, "F")
            return True

        create_pdf_from_charts_comp(charts_for_pdf, pdf_file_path, "TRIAC TIME REPORT - COMPARISON", config_info, logo_path)
        return True
    except Exception as e:
        st.error(f"Lỗi khi tạo báo cáo PDF so sánh: {e}.")
        return False
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)
