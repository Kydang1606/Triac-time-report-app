import pandas as pd
import datetime
import os
import io
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from matplotlib import pyplot as plt
import tempfile
import shutil
import streamlit as st # Import Streamlit for st.error/success
from data_processing_utils import sanitize_filename # Import utility functions

def export_standard_report_excel(df, config, output_file_path, translations):
    """Exports the standard report to an Excel file."""
    mode = config.get('mode', 'year')
    
    groupby_cols = []
    if mode == 'year':
        groupby_cols = ['Year', 'Project name']
    elif mode == 'month':
        groupby_cols = ['Year', 'MonthName', 'Project name']
    else: # week mode
        groupby_cols = ['Year', 'Week', 'Project name']

    if not all(col in df.columns for col in groupby_cols + ['Hours']):
        st.error(translations["error_missing_columns_excel"])
        return False

    if df.empty:
        st.warning(translations["warning_empty_df_excel"])
        return False

    summary = df.groupby(groupby_cols)['Hours'].sum().reset_index()

    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            summary.to_excel(writer, sheet_name=translations["excel_summary_sheet_name"], index=False)

        wb = load_workbook(output_file_path)
        ws = wb[translations["excel_summary_sheet_name"]]
        
        if len(summary) > 0:
            data_col_idx = summary.columns.get_loc('Hours') + 1
            cats_col_idx = summary.columns.get_loc('Project name') + 1

            data_ref = Reference(ws, min_col=data_col_idx, min_row=2, max_row=ws.max_row)
            cats_ref = Reference(ws, min_col=cats_col_idx, min_row=2, max_row=ws.max_row)

            chart = BarChart()
            chart.title = translations["chart_title_hours_by_project"].format(mode=mode)
            chart.x_axis.title = translations["chart_xaxis_project"]
            chart.y_axis.title = translations["chart_yaxis_hours"]
            
            chart.add_data(data_ref, titles_from_data=False) 
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "F2")

        for project in df['Project name'].unique():
            df_proj = df[df['Project name'] == project]
            sheet_title = sanitize_filename(project)
            
            # Check if sheet exists, if so, append data, else create new
            if sheet_title in wb.sheetnames:
                ws_proj = wb[sheet_title]
            else:
                ws_proj = wb.create_sheet(title=sheet_title)

            summary_task = df_proj.groupby('Task')['Hours'].sum().reset_index().sort_values('Hours', ascending=False)
            
            if not summary_task.empty:
                ws_proj.append([translations["task_col_header"], translations["hours_col_header"]])
                for row_data in dataframe_to_rows(summary_task, index=False, header=False):
                    ws_proj.append(row_data)

                chart_task = BarChart()
                chart_task.title = translations["chart_title_hours_by_task"].format(project=project)
                chart_task.x_axis.title = translations["chart_xaxis_task"]
                chart_task.y_axis.title = translations["chart_yaxis_hours"]
                task_len = len(summary_task)
                
                data_ref_task = Reference(ws_proj, min_col=2, min_row=1, max_row=task_len + 1)
                cats_ref_task = Reference(ws_proj, min_col=1, min_row=2, max_row=task_len + 1)
                chart_task.add_data(data_ref_task, titles_from_data=True)
                chart_task.set_categories(cats_ref_task)
                ws_proj.add_chart(chart_task, f"E1")

            start_row_raw_data = ws_proj.max_row + 2 if ws_proj.max_row > 1 else 1
            if not summary_task.empty:
                start_row_raw_data += 15 # Give some space after the chart

            # Add raw data for the project
            # Ensure headers are written only once if appending
            if ws_proj.max_row < start_row_raw_data: # Only write headers if not already written
                for c_idx, header in enumerate(df_proj.columns):
                    ws_proj.cell(row=start_row_raw_data, column=c_idx + 1, value=header)
                start_row_raw_data += 1 # Move to the next row for data

            for r_idx, r in enumerate(dataframe_to_rows(df_proj, index=False, header=False)): # Don't write header again
                for c_idx, cell_val in enumerate(r):
                    ws_proj.cell(row=start_row_raw_data + r_idx, column=c_idx + 1, value=cell_val)
        
        ws_config = wb.create_sheet(translations["excel_config_info_sheet_name"])
        ws_config['A1'], ws_config['B1'] = translations["config_mode_label"], config.get('mode', 'N/A').capitalize()
        ws_config['A2'], ws_config['B2'] = translations["config_year_label"], ', '.join(map(str, config.get('years', []))) if config.get('years') else str(config.get('year', 'N/A'))
        ws_config['A3'], ws_config['B3'] = translations["config_months_label"], ', '.join(config.get('months', [])) if config.get('months') else translations["all_label"]
        
        if 'project_filter_df' in config and not config['project_filter_df'].empty:
            selected_projects_display = config['project_filter_df'][config['project_filter_df']['Include'].astype(str).str.lower() == 'yes']['Project Name'].tolist()
            ws_config['A4'], ws_config['B4'] = translations["config_projects_included_label"], ', '.join(selected_projects_display)
        else:
            ws_config['A4'], ws_config['B4'] = translations["config_projects_included_label"], translations["no_projects_selected_found"]

        # Remove template sheets
        for sheet_name in ['Raw Data', 'Config_Year_Mode', 'Config_Project_Filter']:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        wb.save(output_file_path)
        return True
    except Exception as e:
        st.error(translations["error_export_excel"].format(error=e))
        return False

def export_standard_report_pdf(df, config, pdf_report_path, logo_path, translations):
    """Exports the standard report to a PDF file with charts."""
    today_str = datetime.datetime.today().strftime("%Y-%m-%d")
    tmp_dir = tempfile.mkdtemp()
    charts_for_pdf = []

    def create_pdf_from_charts(charts_data, output_path, title, config_info, logo_path_inner):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font('helvetica', 'B', 16) 
        
        pdf.add_page()
        if os.path.exists(logo_path_inner):
            pdf.image(logo_path_inner, x=10, y=10, w=30)
        pdf.ln(40)
        pdf.cell(0, 10, title, ln=True, align='C')
        pdf.set_font("helvetica", '', 12) 
        pdf.ln(5)
        pdf.cell(0, 10, f"{translations['generated_on_label']}: {today_str}", ln=True, align='C')
        pdf.ln(10)
        pdf.set_font("helvetica", '', 11) 
        for key, value in config_info.items():
            pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')

        for img_path, chart_title, page_project_name in charts_data:
            if img_path and os.path.exists(img_path):
                pdf.add_page()
                if os.path.exists(logo_path_inner):
                    pdf.image(logo_path_inner, x=10, y=8, w=25)
                pdf.set_font("helvetica", 'B', 11) 
                pdf.set_y(35)
                if page_project_name:
                    pdf.cell(0, 10, f"{translations['project_label']}: {page_project_name}", ln=True, align='C')
                pdf.cell(0, 10, chart_title, ln=True, align='C')
                pdf.image(img_path, x=10, y=45, w=190)

        pdf.output(output_path, "F")
        print(f"DEBUG: PDF report generated at {output_path}")

    try:
        projects = df['Project name'].unique() 

        config_info = {
            translations["config_mode_label"]: config.get('mode', 'N/A').capitalize(),
            translations["config_years_label"]: ', '.join(map(str, config.get('years', []))) if config.get('years') else str(config.get('year', 'N/A')),
            translations["config_months_label"]: ', '.join(config.get('months', [])) if config.get('months') else translations["all_label"],
            translations["config_projects_included_label"]: ', '.join(config['project_filter_df']['Project Name']) if 'project_filter_df' in config and not config['project_filter_df'].empty else translations["no_projects_selected_found"]
        }

        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'Liberation Sans']
        plt.rcParams['axes.unicode_minus'] = False 

        for project in projects:
            safe_project = sanitize_filename(project)
            df_proj = df[df['Project name'] == project]

            if 'Workcentre' in df_proj.columns and not df_proj['Workcentre'].empty and df_proj['Workcentre'].value_counts().sum() > 0: # Check for actual data
                workcentre_summary = df_proj.groupby('Workcentre')['Hours'].sum().sort_values(ascending=False)
                if not workcentre_summary.empty and workcentre_summary.sum() > 0:
                    fig, ax = plt.subplots(figsize=(10, 5))
                    workcentre_summary.plot(kind='barh', color='skyblue', ax=ax)
                    ax.set_title(translations["chart_title_hours_by_workcentre"].format(project=project), fontsize=9)
                    ax.tick_params(axis='y', labelsize=8)
                    ax.set_xlabel(translations["chart_xaxis_hours"])
                    ax.set_ylabel(translations["chart_yaxis_workcentre"])
                    wc_img_path = os.path.join(tmp_dir, f"{safe_project}_wc.png")
                    plt.tight_layout()
                    fig.savefig(wc_img_path, dpi=150)
                    plt.close(fig)
                    charts_for_pdf.append((wc_img_path, translations["chart_title_hours_by_workcentre"].format(project=project), project))

            if 'Task' in df_proj.columns and not df_proj['Task'].empty and df_proj['Task'].value_counts().sum() > 0: # Check for actual data
                task_summary = df_proj.groupby('Task')['Hours'].sum().sort_values(ascending=False)
                if not task_summary.empty and task_summary.sum() > 0:
                    fig, ax = plt.subplots(figsize=(10, 6))
                    task_summary.plot(kind='barh', color='lightgreen', ax=ax)
                    ax.set_title(translations["chart_title_hours_by_task"].format(project=project), fontsize=9)
                    ax.tick_params(axis='y', labelsize=8)
                    ax.set_xlabel(translations["chart_xaxis_hours"])
                    ax.set_ylabel(translations["chart_yaxis_task"])
                    task_img_path = os.path.join(tmp_dir, f"{safe_project}_task.png")
                    plt.tight_layout()
                    fig.savefig(task_img_path, dpi=150)
                    plt.close(fig)
                    charts_for_pdf.append((task_img_path, translations["chart_title_hours_by_task"].format(project=project), project))

        if not charts_for_pdf:
            st.warning(translations["warning_no_charts_pdf"])
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, translations["pdf_report_title_standard"], ln=True, align='C')
            pdf.set_font("helvetica", '', 12)
            pdf.cell(0, 10, f"{translations['generated_on_label']}: {today_str}", ln=True, align='C')
            pdf.ln(10)
            pdf.set_font("helvetica", '', 11)
            for key, value in config_info.items():
                pdf.cell(0, 7, f"{key}: {value}", ln=True, align='C')
            pdf.cell(0, 10, translations["no_charts_generated_message"], ln=True, align='C')
            pdf.output(pdf_report_path, "F")
            return True
            
        create_pdf_from_charts(charts_for_pdf, pdf_report_path, translations["pdf_report_title_standard"], config_info, logo_path)
        return True
    except Exception as e:
        st.error(translations["error_create_pdf"].format(error=e))
        return False
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)
